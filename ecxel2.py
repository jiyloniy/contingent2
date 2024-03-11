import os
import re
import os
import re

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()
from user.models import Faculty, Budjet, Shartnoma, Organization, Yonalish, Guruh

from datetime import datetime

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from openpyxl.utils import get_column_letter


def set_cell_properties(cell, value, alignment, font, border):
    cell.value = value
    cell.alignment = alignment
    cell.font = font
    cell.border = border


import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()
from user.models import Faculty, Budjet, Shartnoma, Organization, Yonalish, Guruh

org = Organization.objects.filter(name='kiuf').first()

FONT_NAME = 'Times New Roman'
FONT_SIZE = 8
FONT_COLOR = 'FF000000'
BORDER_STYLE = 'thin'
BORDER_COLOR = 'FF000000'
red_color = 'FF0000FF'
# color blue
blue_color = 'FFFF0000'
wb = Workbook()
ws = wb.active
# add auto size width of column

# Hozirgi vaqtni olib, formatini belgilash

now = datetime.now()
formatted_time = now.strftime("%Y-%m-%d")


def exporttoexcel(org):
    kurss = Guruh.objects.filter(org=org).values('kurs').distinct()
    organization_name = org.full_name
    ws.merge_cells('A1:AD2')
    set_cell_properties(ws.cell(row=1, column=1),
                        f"{organization_name} talabalari kontingentining {formatted_time} holati haqida umumiy ma'lumot (o'zbek /rus)",
                        Alignment(horizontal='center', vertical='center'),
                        Font(name=FONT_NAME, size=18, bold=True, italic=False, color=FONT_COLOR),
                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

    cell_properties = [
        {'row': 3, 'column': 1, 'value': '№', 'width': True, 'merge': 'A3:A4'},
        {'row': 3, 'column': 2, 'value': 'Ta\'lim yo\'nalishi kodi va nomi', 'width': True, 'merge': 'B3:B4'},
        {'row': 3, 'column': 3, 'value': 'Ta\'lim turi', 'width': True, 'merge': 'C3:C4'},
        {'row': 3, 'column': 4, 'value': 'Jami', 'width': True, 'merge': 'D3:D4'},
        {'row': 3, 'column': 5, 'value': 'Jami', 'width': True, 'merge': 'E3:F3'},
        {'row': 4, 'column': 5, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 6, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 7, 'value': '1-kurs', 'width': True, 'merge': 'G3:I3'},
        {'row': 4, 'column': 7, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 8, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 9, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 10, 'value': '2-kurs', 'width': True, 'merge': 'J3:L3'},
        {'row': 4, 'column': 10, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 11, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 12, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 13, 'value': '3-kurs', 'width': True, 'merge': 'M3:O3'},
        {'row': 4, 'column': 13, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 14, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 15, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 16, 'value': '4-kurs', 'width': True, 'merge': 'P3:R3'},
        {'row': 4, 'column': 16, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 17, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 18, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 19, 'value': '4-kurs', 'width': True, 'merge': 'S3:U3'},
        {'row': 4, 'column': 19, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 20, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 21, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 22, 'value': '5-kurs', 'width': True, 'merge': 'V3:X3'},
        {'row': 4, 'column': 22, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 23, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 24, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 25, 'value': '5-kurs', 'width': True, 'merge': 'Y3:AA3'},
        {'row': 4, 'column': 25, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 26, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 27, 'value': 'Rus', 'width': True},
        {'row': 3, 'column': 28, 'value': '6-kurs', 'width': True, 'merge': 'AB3:AD3'},
        {'row': 4, 'column': 28, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 29, 'value': 'O\'zbek', 'width': True},
        {'row': 4, 'column': 30, 'value': 'Rus', 'width': True},
    ]

    for properties in cell_properties:
        cell = ws.cell(row=properties['row'], column=properties['column'])
        set_cell_properties(cell,
                            properties['value'],
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
        # if 'width' in properties and properties['width']:
        #     ws.column_dimensions[get_column_letter(properties['column'])].auto_size = True
        if 'merge' in properties:
            ws.merge_cells(properties['merge'])

    yonalish_kunduzgi = Yonalish.objects.filter(org=org, turi='Kunduzgi', mutahasislik_2=False).exclude(
        yonalishguruh__bosqich='Magistr')
    yonalish_kunduzgi = yonalish_kunduzgi.exclude(mutahasislik_2=True)
    yonalish_sirtqi = Yonalish.objects.filter(org=org, turi='Sirtqi', mutahasislik_2=False).exclude(
        yonalishguruh__bosqich='Magistr')
    yonalish_masofaviy = Yonalish.objects.filter(org=org, turi='Masofaviy', mutahasislik_2=False).exclude(
        yonalishguruh__bosqich='Magistr')
    mut_2 = Yonalish.objects.filter(org=org, mutahasislik_2=True).exclude(
        yonalishguruh__bosqich='Magistr', turi='Masofaviy')
    magistir = Yonalish.objects.filter(org=org, yonalishguruh__bosqich='Magistr')

    print(mut_2)
    kurs_jami = 0

    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    row = 5
    for kunduzgi_yonlaish in yonalish_kunduzgi:
        jami = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)
            if kunduzgi_yonlaish.language == 'O\'zbek':
                ws.cell(row=row, column=5, value=jami)
                ws.cell(row=row, column=6, value=0)
                jami_uzek += jami
            else:
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=jami)


            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'))
                    if budget['jami']:
                        jami += budget['jami']
                    if shartnoma['jami']:
                        jami += shartnoma['jami']

                    if k == 1 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami)
                        ws.cell(row=row, column=9, value=0)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += jami
                        kurs_1_rus_jami += 0
                    elif k == 1 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=0)
                        ws.cell(row=row, column=9, value=jami)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += 0
                        kurs_1_rus_jami += jami
                    elif k == 2 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami)
                        ws.cell(row=row, column=12, value=0)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += jami
                        kurs_2_rus_jami += 0
                    elif k == 2 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=0)
                        ws.cell(row=row, column=12, value=jami)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += 0
                        kurs_2_rus_jami += jami
                    elif k == 3 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami)
                        ws.cell(row=row, column=15, value=0)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += jami
                        kurs_3_rus_jami += 0
                    elif k == 3 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=0)
                        ws.cell(row=row, column=15, value=jami)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += 0
                        kurs_3_rus_jami += jami
                    elif k == 4 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami)
                        ws.cell(row=row, column=18, value=0)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += jami
                        kurs_4_rus_jami += 0
                    elif k == 4 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=0)
                        ws.cell(row=row, column=18, value=jami)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += 0
                        kurs_4_rus_jami += jami
                    elif k == 5 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami)
                        ws.cell(row=row, column=21, value=0)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += jami
                        kurs_5_rus_jami += 0
                    elif k == 5 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=0)
                        ws.cell(row=row, column=21, value=jami)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += 0
                        kurs_5_rus_jami += jami
                    elif k == 6 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami)
                        ws.cell(row=row, column=24, value=0)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += jami
                        kurs_6_rus_jami += 0
                    elif k == 6 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=0)
                        ws.cell(row=row, column=24, value=jami)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=0)
                        ws.cell(row=row, column=27, value=jami)

                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)

            row += 1

            ws.cell(row=row, column=2, value='Kunduzgi Jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_uzek)
            ws.cell(row=row, column=6, value=jami_rus)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
            ws.cell(row=row, column=9, value=kurs_1_rus_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
            ws.cell(row=row, column=12, value=kurs_2_rus_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
            ws.cell(row=row, column=15, value=kurs_3_rus_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
            ws.cell(row=row, column=18, value=kurs_4_rus_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
            ws.cell(row=row, column=21, value=kurs_5_rus_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
            ws.cell(row=row, column=24, value=kurs_6_rus_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
            ws.cell(row=row, column=27, value=kurs_7_rus_jami)


    row += 1
    jami_full = 0
    kurs_jami = 0
    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    for kunduzgi_yonlaish in yonalish_sirtqi:
        jami = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)
            if kunduzgi_yonlaish.language == 'O\'zbek':
                ws.cell(row=row, column=5, value=jami)
                ws.cell(row=row, column=6, value=0)
                jami_uzek += jami
            else:
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=jami)
                jami_rus += jami

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'))
                    if budget['jami']:
                        jami += budget['jami']
                    if shartnoma['jami']:
                        jami += shartnoma['jami']

                    if k == 1 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami)
                        ws.cell(row=row, column=9, value=0)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += jami
                        kurs_1_rus_jami += 0
                    elif k == 1 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=0)
                        ws.cell(row=row, column=9, value=jami)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += 0
                        kurs_1_rus_jami += jami
                    elif k == 2 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami)
                        ws.cell(row=row, column=12, value=0)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += jami
                        kurs_2_rus_jami += 0
                    elif k == 2 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=0)
                        ws.cell(row=row, column=12, value=jami)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += 0
                        kurs_2_rus_jami += jami
                    elif k == 3 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami)
                        ws.cell(row=row, column=15, value=0)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += jami
                        kurs_3_rus_jami += 0
                    elif k == 3 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=0)
                        ws.cell(row=row, column=15, value=jami)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += 0
                        kurs_3_rus_jami += jami
                    elif k == 4 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami)
                        ws.cell(row=row, column=18, value=0)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += jami
                        kurs_4_rus_jami += 0
                    elif k == 4 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=0)
                        ws.cell(row=row, column=18, value=jami)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += 0
                        kurs_4_rus_jami += jami
                    elif k == 5 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami)
                        ws.cell(row=row, column=21, value=0)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += jami
                        kurs_5_rus_jami += 0
                    elif k == 5 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=0)
                        ws.cell(row=row, column=21, value=jami)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += 0
                        kurs_5_rus_jami += jami
                    elif k == 6 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami)
                        ws.cell(row=row, column=24, value=0)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += jami
                        kurs_6_rus_jami += 0
                    elif k == 6 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=0)
                        ws.cell(row=row, column=24, value=jami)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=0)
                        ws.cell(row=row, column=27, value=jami)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)

            row += 1

            ws.cell(row=row, column=2, value='Sirtqi Jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_uzek)
            ws.cell(row=row, column=6, value=jami_rus)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
            ws.cell(row=row, column=9, value=kurs_1_rus_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
            ws.cell(row=row, column=12, value=kurs_2_rus_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
            ws.cell(row=row, column=15, value=kurs_3_rus_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
            ws.cell(row=row, column=18, value=kurs_4_rus_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
            ws.cell(row=row, column=21, value=kurs_5_rus_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
            ws.cell(row=row, column=24, value=kurs_6_rus_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
            ws.cell(row=row, column=27, value=kurs_7_rus_jami)
    row += 1
    jami_full = 0
    kurs_jami = 0
    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    for kunduzgi_yonlaish in yonalish_masofaviy:
        jami = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)
            if kunduzgi_yonlaish.language == 'O\'zbek':
                ws.cell(row=row, column=5, value=jami)
                ws.cell(row=row, column=6, value=0)
                jami_uzek += jami
            else:
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=jami)


            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'))
                    if budget['jami']:
                        jami += budget['jami']
                    if shartnoma['jami']:
                        jami += shartnoma['jami']

                    if k == 1 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami)
                        ws.cell(row=row, column=9, value=0)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += jami
                        kurs_1_rus_jami += 0
                    elif k == 1 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=0)
                        ws.cell(row=row, column=9, value=jami)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += 0
                        kurs_1_rus_jami += jami
                    elif k == 2 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami)
                        ws.cell(row=row, column=12, value=0)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += jami
                        kurs_2_rus_jami += 0
                    elif k == 2 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=0)
                        ws.cell(row=row, column=12, value=jami)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += 0
                        kurs_2_rus_jami += jami
                    elif k == 3 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami)
                        ws.cell(row=row, column=15, value=0)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += jami
                        kurs_3_rus_jami += 0
                    elif k == 3 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=0)
                        ws.cell(row=row, column=15, value=jami)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += 0
                        kurs_3_rus_jami += jami
                    elif k == 4 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami)
                        ws.cell(row=row, column=18, value=0)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += jami
                        kurs_4_rus_jami += 0
                    elif k == 4 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=0)
                        ws.cell(row=row, column=18, value=jami)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += 0
                        kurs_4_rus_jami += jami
                    elif k == 5 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami)
                        ws.cell(row=row, column=21, value=0)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += jami
                        kurs_5_rus_jami += 0
                    elif k == 5 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=0)
                        ws.cell(row=row, column=21, value=jami)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += 0
                        kurs_5_rus_jami += jami
                    elif k == 6 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami)
                        ws.cell(row=row, column=24, value=0)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += jami
                        kurs_6_rus_jami += 0
                    elif k == 6 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=0)
                        ws.cell(row=row, column=24, value=jami)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=0)
                        ws.cell(row=row, column=27, value=jami)

                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)

            row += 1

            ws.cell(row=row, column=2, value='Masofaviy jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_uzek)
            ws.cell(row=row, column=6, value=jami_rus)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
            ws.cell(row=row, column=9, value=kurs_1_rus_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
            ws.cell(row=row, column=12, value=kurs_2_rus_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
            ws.cell(row=row, column=15, value=kurs_3_rus_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
            ws.cell(row=row, column=18, value=kurs_4_rus_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
            ws.cell(row=row, column=21, value=kurs_5_rus_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
            ws.cell(row=row, column=24, value=kurs_6_rus_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
            ws.cell(row=row, column=27, value=kurs_7_rus_jami)
    row += 1
    jami_full = 0
    kurs_jami = 0
    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    for kunduzgi_yonlaish in mut_2:
        jami = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)
            if kunduzgi_yonlaish.language == 'O\'zbek':
                ws.cell(row=row, column=5, value=jami)
                ws.cell(row=row, column=6, value=0)
                jami_uzek += jami
            else:
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=jami)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'))
                    if budget['jami']:
                        jami += budget['jami']
                    if shartnoma['jami']:
                        jami += shartnoma['jami']

                    if k == 1 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami)
                        ws.cell(row=row, column=9, value=0)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += jami
                        kurs_1_rus_jami += 0
                    elif k == 1 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=0)
                        ws.cell(row=row, column=9, value=jami)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += 0
                        kurs_1_rus_jami += jami
                    elif k == 2 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami)
                        ws.cell(row=row, column=12, value=0)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += jami
                        kurs_2_rus_jami += 0
                    elif k == 2 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=0)
                        ws.cell(row=row, column=12, value=jami)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += 0
                        kurs_2_rus_jami += jami
                    elif k == 3 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami)
                        ws.cell(row=row, column=15, value=0)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += jami
                        kurs_3_rus_jami += 0
                    elif k == 3 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=0)
                        ws.cell(row=row, column=15, value=jami)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += 0
                        kurs_3_rus_jami += jami
                    elif k == 4 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami)
                        ws.cell(row=row, column=18, value=0)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += jami
                        kurs_4_rus_jami += 0
                    elif k == 4 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=0)
                        ws.cell(row=row, column=18, value=jami)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += 0
                        kurs_4_rus_jami += jami
                    elif k == 5 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami)
                        ws.cell(row=row, column=21, value=0)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += jami
                        kurs_5_rus_jami += 0
                    elif k == 5 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=0)
                        ws.cell(row=row, column=21, value=jami)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += 0
                        kurs_5_rus_jami += jami
                    elif k == 6 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami)
                        ws.cell(row=row, column=24, value=0)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += jami
                        kurs_6_rus_jami += 0
                    elif k == 6 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=0)
                        ws.cell(row=row, column=24, value=jami)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=0)
                        ws.cell(row=row, column=27, value=jami)

                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)

            row += 1

            ws.cell(row=row, column=2, value='Masofaviy jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_uzek)
            ws.cell(row=row, column=6, value=jami_rus)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
            ws.cell(row=row, column=9, value=kurs_1_rus_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
            ws.cell(row=row, column=12, value=kurs_2_rus_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
            ws.cell(row=row, column=15, value=kurs_3_rus_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
            ws.cell(row=row, column=18, value=kurs_4_rus_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
            ws.cell(row=row, column=21, value=kurs_5_rus_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
            ws.cell(row=row, column=24, value=kurs_6_rus_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
            ws.cell(row=row, column=27, value=kurs_7_rus_jami)
    row += 1
    jami_full = 0
    kurs_jami = 0
    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    for kunduzgi_yonlaish in magistir:
        jami = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)
            if kunduzgi_yonlaish.language == 'O\'zbek':
                ws.cell(row=row, column=5, value=jami)
                ws.cell(row=row, column=6, value=0)
                jami_uzek += jami
            else:
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=jami)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'))
                    if budget['jami']:
                        jami += budget['jami']
                    if shartnoma['jami']:
                        jami += shartnoma['jami']

                    if k == 1 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami)
                        ws.cell(row=row, column=9, value=0)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += jami
                        kurs_1_rus_jami += 0
                    elif k == 1 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=0)
                        ws.cell(row=row, column=9, value=jami)
                        kurs_1_jami += jami
                        kurs_1_uzbek_jami += 0
                        kurs_1_rus_jami += jami
                    elif k == 2 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami)
                        ws.cell(row=row, column=12, value=0)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += jami
                        kurs_2_rus_jami += 0
                    elif k == 2 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=0)
                        ws.cell(row=row, column=12, value=jami)
                        kurs_2_jami += jami
                        kurs_2_uzbek_jami += 0
                        kurs_2_rus_jami += jami
                    elif k == 3 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami)
                        ws.cell(row=row, column=15, value=0)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += jami
                        kurs_3_rus_jami += 0
                    elif k == 3 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=0)
                        ws.cell(row=row, column=15, value=jami)
                        kurs_3_jami += jami
                        kurs_3_uzbek_jami += 0
                        kurs_3_rus_jami += jami
                    elif k == 4 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami)
                        ws.cell(row=row, column=18, value=0)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += jami
                        kurs_4_rus_jami += 0
                    elif k == 4 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=0)
                        ws.cell(row=row, column=18, value=jami)
                        kurs_4_jami += jami
                        kurs_4_uzbek_jami += 0
                        kurs_4_rus_jami += jami
                    elif k == 5 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami)
                        ws.cell(row=row, column=21, value=0)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += jami
                        kurs_5_rus_jami += 0
                    elif k == 5 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=0)
                        ws.cell(row=row, column=21, value=jami)
                        kurs_5_jami += jami
                        kurs_5_uzbek_jami += 0
                        kurs_5_rus_jami += jami
                    elif k == 6 and kunduzgi_yonlaish.language == 'O\'zbek':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami)
                        ws.cell(row=row, column=24, value=0)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += jami
                        kurs_6_rus_jami += 0
                    elif k == 6 and kunduzgi_yonlaish.language == 'Rus':

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=0)
                        ws.cell(row=row, column=24, value=jami)
                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=0)
                        ws.cell(row=row, column=27, value=jami)

                        kurs_6_jami += jami
                        kurs_6_uzbek_jami += 0
                        kurs_6_rus_jami += jami
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)

            row += 1

            ws.cell(row=row, column=2, value='Magistir jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_uzek)
            ws.cell(row=row, column=6, value=jami_rus)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
            ws.cell(row=row, column=9, value=kurs_1_rus_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
            ws.cell(row=row, column=12, value=kurs_2_rus_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
            ws.cell(row=row, column=15, value=kurs_3_rus_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
            ws.cell(row=row, column=18, value=kurs_4_rus_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
            ws.cell(row=row, column=21, value=kurs_5_rus_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
            ws.cell(row=row, column=24, value=kurs_6_rus_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
            ws.cell(row=row, column=27, value=kurs_7_rus_jami)
    row += 1
    # bakalvr bo'gan guruhlar jamini hisoblash
    jami_full = 0
    kurs_jami = 0
    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    result = {}
    guruhs = Guruh.objects.filter(org=org, bosqich='Bakalavr')
    for guruh in guruhs:
        language = guruh.yonalish.language
        course = guruh.kurs
        jami = 0
        budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
        shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
    for k in [1, 2, 3, 4, 5, 6]:
        guruhs = Guruh.objects.filter(org=org, bosqich='Bakalavr', kurs=k)
        for guruh in guruhs:
            language = guruh.yonalish.language
            course = guruh.kurs
            jami = 0
            budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
            shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
            if budget['jami']:
                jami += budget['jami']
            if shartnoma['jami']:
                jami += shartnoma['jami']
            if language == 'O\'zbek':
                jami_uzek += jami
            else:
                jami_rus += jami
            if k == 1:
                if language == 'O\'zbek':
                    kurs_1_uzbek_jami += jami
                else:
                    kurs_1_rus_jami += jami
                kurs_1_jami += jami
            elif k == 2:
                if language == 'O\'zbek':
                    kurs_2_uzbek_jami += jami
                else:
                    kurs_2_rus_jami += jami
                kurs_2_jami += jami
            elif k == 3:
                if language == 'O\'zbek':
                    kurs_3_uzbek_jami += jami
                else:
                    kurs_3_rus_jami += jami
                kurs_3_jami += jami
            elif k == 4:
                if language == 'O\'zbek':
                    kurs_4_uzbek_jami += jami
                else:
                    kurs_4_rus_jami += jami
                kurs_4_jami += jami
            elif k == 5:
                if language == 'O\'zbek':
                    kurs_5_uzbek_jami += jami
                else:
                    kurs_5_rus_jami += jami
                kurs_5_jami += jami
            elif k == 6:
                if language == 'O\'zbek':
                    kurs_6_uzbek_jami += jami
                else:
                    kurs_6_rus_jami += jami
                kurs_6_jami += jami
            else:
                if language == 'O\'zbek':
                    kurs_7_uzbek_jami+=jami

                else:
                    kurs_7_rus_jami+=jami
                kurs_7_jami += jami
    row += 1
    ws.cell(row=row, column=2, value='Bakalavr full')
    ws.cell(row=row, column=4, value=jami_full)
    ws.cell(row=row, column=5, value=jami_uzek)
    ws.cell(row=row, column=6, value=jami_rus)
    ws.cell(row=row, column=7, value=kurs_1_jami)
    ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
    ws.cell(row=row, column=9, value=kurs_1_rus_jami)
    ws.cell(row=row, column=10, value=kurs_2_jami)
    ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
    ws.cell(row=row, column=12, value=kurs_2_rus_jami)
    ws.cell(row=row, column=13, value=kurs_3_jami)
    ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
    ws.cell(row=row, column=15, value=kurs_3_rus_jami)
    ws.cell(row=row, column=16, value=kurs_4_jami)
    ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
    ws.cell(row=row, column=18, value=kurs_4_rus_jami)
    ws.cell(row=row, column=19, value=kurs_5_jami)
    ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
    ws.cell(row=row, column=21, value=kurs_5_rus_jami)
    ws.cell(row=row, column=22, value=kurs_6_jami)
    ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
    ws.cell(row=row, column=24, value=kurs_6_rus_jami)
    ws.cell(row=row, column=25, value=kurs_7_jami)
    ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
    ws.cell(row=row, column=27, value=kurs_7_rus_jami)
    row+=1
    jami_full = 0
    kurs_jami = 0
    jami_full = 0
    jami_uzek = 0
    jami_rus = 0
    kurs_1_jami = 0
    kurs_1_uzbek_jami = 0
    kurs_1_rus_jami = 0
    kurs_2_jami = 0
    kurs_2_uzbek_jami = 0
    kurs_2_rus_jami = 0
    kurs_3_jami = 0
    kurs_3_uzbek_jami = 0
    kurs_3_rus_jami = 0
    kurs_4_jami = 0
    kurs_4_uzbek_jami = 0
    kurs_4_rus_jami = 0
    kurs_5_jami = 0
    kurs_5_uzbek_jami = 0
    kurs_5_rus_jami = 0
    kurs_6_jami = 0
    kurs_6_uzbek_jami = 0
    kurs_6_rus_jami = 0
    kurs_7_jami = 0
    kurs_7_uzbek_jami = 0
    kurs_7_rus_jami = 0
    fakultetlar = Faculty.objects.filter(org=org)
    for fakultet in fakultetlar:

        guruhs = Guruh.objects.filter(yonalish__faculty=fakultet)
        for guruh in guruhs:
            language = guruh.yonalish.language
            course = guruh.kurs
            jami = 0
            budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
            shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
            if budget['jami']:
                jami += budget['jami']
                jami_full += budget['jami']
            if shartnoma['jami']:
                jami += shartnoma['jami']
                jami_full += shartnoma['jami']
        for k in [1, 2, 3, 4, 5, 6]:
            guruhs = Guruh.objects.filter(org=org, bosqich='Bakalavr', kurs=k)
            for guruh in guruhs:
                language = guruh.yonalish.language
                course = guruh.kurs
                jami = 0
                budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
                shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
                if budget['jami']:
                    jami += budget['jami']
                if shartnoma['jami']:
                    jami += shartnoma['jami']
                if language == 'O\'zbek':
                    jami_uzek += jami
                else:
                    jami_rus += jami
                if k == 1:
                    if language == 'O\'zbek':
                        kurs_1_uzbek_jami += jami
                    else:
                        kurs_1_rus_jami += jami
                    kurs_1_jami += jami
                elif k == 2:
                    if language == 'O\'zbek':
                        kurs_2_uzbek_jami += jami
                    else:
                        kurs_2_rus_jami += jami
                    kurs_2_jami += jami
                elif k == 3:
                    if language == 'O\'zbek':
                        kurs_3_uzbek_jami += jami
                    else:
                        kurs_3_rus_jami += jami
                    kurs_3_jami += jami
                elif k == 4:
                    if language == 'O\'zbek':
                        kurs_4_uzbek_jami += jami
                    else:
                        kurs_4_rus_jami += jami
                    kurs_4_jami += jami
                elif k == 5:
                    if language == 'O\'zbek':
                        kurs_5_uzbek_jami += jami
                    else:
                        kurs_5_rus_jami += jami
                    kurs_5_jami += jami
                elif k == 6:
                    if language == 'O\'zbek':
                        kurs_6_uzbek_jami += jami
                    else:
                        kurs_6_rus_jami += jami
                    kurs_6_jami += jami
                else:
                    if language == 'O\'zbek':
                        kurs_7_uzbek_jami += jami

                    else:
                        kurs_7_rus_jami += jami
                    kurs_7_jami += jami
        row += 1
        ws.cell(row=row, column=2, value=f'{fakultet.name}')
        ws.cell(row=row, column=4, value=jami_full)
        ws.cell(row=row, column=5, value=jami_uzek)
        ws.cell(row=row, column=6, value=jami_rus)
        ws.cell(row=row, column=7, value=kurs_1_jami)
        ws.cell(row=row, column=8, value=kurs_1_uzbek_jami)
        ws.cell(row=row, column=9, value=kurs_1_rus_jami)
        ws.cell(row=row, column=10, value=kurs_2_jami)
        ws.cell(row=row, column=11, value=kurs_2_uzbek_jami)
        ws.cell(row=row, column=12, value=kurs_2_rus_jami)
        ws.cell(row=row, column=13, value=kurs_3_jami)
        ws.cell(row=row, column=14, value=kurs_3_uzbek_jami)
        ws.cell(row=row, column=15, value=kurs_3_rus_jami)
        ws.cell(row=row, column=16, value=kurs_4_jami)
        ws.cell(row=row, column=17, value=kurs_4_uzbek_jami)
        ws.cell(row=row, column=18, value=kurs_4_rus_jami)
        ws.cell(row=row, column=19, value=kurs_5_jami)
        ws.cell(row=row, column=20, value=kurs_5_uzbek_jami)
        ws.cell(row=row, column=21, value=kurs_5_rus_jami)
        ws.cell(row=row, column=22, value=kurs_6_jami)
        ws.cell(row=row, column=23, value=kurs_6_uzbek_jami)
        ws.cell(row=row, column=24, value=kurs_6_rus_jami)
        ws.cell(row=row, column=25, value=kurs_7_jami)
        ws.cell(row=row, column=26, value=kurs_7_uzbek_jami)
        ws.cell(row=row, column=27, value=kurs_7_rus_jami)
    row+=1
    guruhs = Guruh.objects.filter(org=org)





    wb.save('talabalar.xlsx')


exporttoexcel(org)
