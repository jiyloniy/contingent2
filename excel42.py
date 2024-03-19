import os

from io import BytesIO

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()

from datetime import datetime

from django.db.models import Sum

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()
from user.models import Faculty, Budjet, Shartnoma, Organization, Yonalish, Guruh




def exporttoexcel4(org):
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl import Workbook

    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")

    def set_cell_properties(cell, value, alignment, font, border):
        cell.value = value
        cell.alignment = alignment
        cell.font = font
        cell.border = border

    FONT_NAME = 'Times New Roman'
    FONT_SIZE = 8
    FONT_COLOR = 'FF000000'
    BORDER_STYLE = 'thin'
    BORDER_COLOR = 'FF000000'
    red_color = 'FF0000FF'
    # color blue
    blue_color = 'FFFF0000'

    # add auto size width of column

    # Hozirgi vaqtni ol
    # add auto size width of column

    # Hozirgi vaqtni olib, formatini belgilash

    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")
    wb = Workbook()
    ws = wb.active
    output4 = BytesIO()
    organization_name = org.full_name

    ws.merge_cells('A1:AD2')
    set_cell_properties(ws.cell(row=1, column=1),
                        f"{organization_name} talabalari kontingentining {formatted_time} holati haqida umumiy ma'lumot (Davlat granti/To'lov-kontrakt)",
                        Alignment(horizontal='center', vertical='center'),
                        Font(name=FONT_NAME, size=18, bold=True, italic=False, color=FONT_COLOR),
                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

    cell_properties = [{'row': 3, 'column': 1, 'value': '№', 'width': True, 'merge': 'A3:A4'},
                       {'row': 3, 'column': 2, 'value': 'Ta\'lim yo\'nalishi kodi va nomi', 'width': True,
                        'merge': 'B3:B4'},
                       {'row': 3, 'column': 3, 'value': 'Ta\'lim turi', 'width': True, 'merge': 'C3:C4'},
                       {'row': 3, 'column': 4, 'value': 'Jami', 'width': True, 'merge': 'D3:D4'},
                       {'row': 3, 'column': 5, 'value': 'Jami', 'width': True, 'merge': 'E3:F3'},
                       {'row': 4, 'column': 5, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 6, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 7, 'value': '1-kurs', 'width': True, 'merge': 'G3:I3'},
                       {'row': 4, 'column': 7, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 8, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 9, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 10, 'value': '2-kurs', 'width': True, 'merge': 'J3:L3'},
                       {'row': 4, 'column': 10, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 11, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 12, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 13, 'value': '3-kurs', 'width': True, 'merge': 'M3:O3'},
                       {'row': 4, 'column': 13, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 14, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 15, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 16, 'value': '4-kurs', 'width': True, 'merge': 'P3:R3'},
                       {'row': 4, 'column': 16, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 17, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 18, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 19, 'value': '4-kurs', 'width': True, 'merge': 'S3:U3'},
                       {'row': 4, 'column': 19, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 20, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 21, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 22, 'value': '5-kurs', 'width': True, 'merge': 'V3:X3'},
                       {'row': 4, 'column': 22, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 23, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 24, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 25, 'value': '5-kurs', 'width': True, 'merge': 'Y3:AA3'},
                       {'row': 4, 'column': 25, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 26, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 27, 'value': 'Ayol', 'width': True},
                       {'row': 3, 'column': 28, 'value': '6-kurs', 'width': True, 'merge': 'AB3:AD3'},
                       {'row': 4, 'column': 28, 'value': 'Jami', 'width': True},
                       {'row': 4, 'column': 29, 'value': 'Erkak', 'width': True},
                       {'row': 4, 'column': 30, 'value': 'Ayol', 'width': True},
                       ]

    for properties in cell_properties:
        cell = ws.cell(row=properties['row'], column=properties['column'])
        set_cell_properties(cell,
                            properties['value'],
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
        # if 'width' in properties and properties['width']:
        #     ws.column_dimensions[get_column_letter(properties['column'])].auto_size = True
        if 'merge' in properties:
            ws.merge_cells(properties['merge'])

    yonalish_kunduzgi = Yonalish.objects.filter(org=org, turi='Kunduzgi', mutahasislik_2=False).exclude(
        yonalishguruh__bosqich='Magistr')
    yonalish_kunduzgi = yonalish_kunduzgi.exclude(mutahasislik_2=True)
    yonalish_sirtqi = Yonalish.objects.filter(org=org, turi='Sirtqi', mutahasislik_2=False).exclude(
        yonalishguruh__bosqich='Magistr')
    yonalish_masofaviy = Yonalish.objects.filter(org=org, turi='Masofaviy', mutahasislik_2=False).exclude(
        yonalishguruh__bosqich='Magistr')
    mut_2 = Yonalish.objects.filter(org=org, mutahasislik_2=True).exclude(
        yonalishguruh__bosqich='Magistr', turi='Masofaviy')
    magistir = Yonalish.objects.filter(org=org, yonalishguruh__bosqich='Magistr')



    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0

    row = 5
    for kunduzgi_yonlaish in yonalish_kunduzgi:
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        budget2 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhbudjet__xotin_qiz'))
        budget3 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhshartnoma__xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']

                jami_erkak += budget['jami'] - budget2['ayollar']
                jami_full_ayol += budget2['ayollar']
                jami_full_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']

                jami_full_ayol += budget3['ayollar']
                jami_full_erkak += shartnoma['jami'] - budget3['ayollar']

            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)

            ws.cell(row=row, column=5, value=jami_ayol)
            ws.cell(row=row, column=6, value=jami_erkak)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    jami_erkak = 0
                    jami_ayol = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'), ayollar=Sum('xotin_qiz'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'),
                                                                                  ayollar=Sum('xotin_qiz'))
                    budget3 = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                       guruh__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    budget2 = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                    guruhi__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    if budget['jami']:
                        jami += budget['jami']
                        if budget2['ayollar']:
                            jami_ayol += budget2['ayollar']
                            jami_erkak += budget['jami'] - budget2['ayollar']
                        else:
                            jami_erkak += budget['jami']
                            jami_full_erkak += budget['jami']

                    if shartnoma['jami']:
                        jami += shartnoma['jami']
                        if budget3['ayollar']:
                            jami_ayol += budget3['ayollar']
                            jami_erkak += shartnoma['jami'] - budget3['ayollar']
                        else:
                            jami_erkak += shartnoma['jami']
                            jami_full_erkak += shartnoma['jami']

                    if k == 1:

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami_erkak)
                        ws.cell(row=row, column=9, value=jami_ayol)
                        kurs_1_jami += jami
                        kurs_1_erkak_jami += jami_erkak
                        kurs_1_ayol_jami += jami_ayol
                    elif k == 2:
                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami_erkak)
                        ws.cell(row=row, column=12, value=jami_ayol)
                        kurs_2_jami += jami
                        kurs_2_erkak_jami += jami_erkak
                        kurs_2_ayol_jami += jami_ayol
                    elif k == 3:

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami_erkak)
                        ws.cell(row=row, column=15, value=jami_ayol)
                        kurs_3_jami += jami
                        kurs_3_erkak_jami += jami_erkak
                        kurs_3_ayol_jami += jami_ayol
                    elif k == 4:

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami_erkak)
                        ws.cell(row=row, column=18, value=jami_ayol)
                        kurs_4_jami += jami
                        kurs_4_erkak_jami += jami_erkak
                        kurs_4_ayol_jami += jami_ayol
                    elif k == 5:

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami_erkak)
                        ws.cell(row=row, column=21, value=jami_ayol)
                        kurs_5_jami += jami
                        kurs_5_erkak_jami += jami_erkak
                        kurs_5_ayol_jami += jami_ayol
                    elif k == 6:

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami_erkak)
                        ws.cell(row=row, column=24, value=jami_ayol)
                        kurs_6_jami += jami
                        kurs_6_erkak_jami += jami_erkak
                        kurs_6_ayol_jami += jami_ayol
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=jami_erkak)
                        ws.cell(row=row, column=27, value=jami_ayol)
                        kurs_7_jami += jami
                        kurs_7_erkak_jami += jami_erkak
                        kurs_7_ayol_jami += jami_ayol
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)
            row += 1

            ws.cell(row=row, column=2, value='Kunduzgi Jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_full_ayol)
            ws.cell(row=row, column=6, value=jami_full_erkak)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
            ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
            ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
            ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
            ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
            ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
            ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
            ws.cell(row=row, column=27, value=kurs_7_ayol_jami)

    row += 1
    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0

    for kunduzgi_yonlaish in yonalish_sirtqi:

        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        budget2 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhbudjet__xotin_qiz'))
        budget3 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhshartnoma__xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']

                jami_erkak += budget['jami'] - budget2['ayollar']
                jami_full_ayol += budget2['ayollar']
                jami_full_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']

                jami_full_ayol += budget3['ayollar']
                jami_full_erkak += shartnoma['jami'] - budget3['ayollar']

            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)

            ws.cell(row=row, column=5, value=jami_ayol)
            ws.cell(row=row, column=6, value=jami_erkak)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    jami_erkak = 0
                    jami_ayol = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'), ayollar=Sum('xotin_qiz'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'),
                                                                                  ayollar=Sum('xotin_qiz'))
                    budget3 = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                       guruh__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    budget2 = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                    guruhi__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    if budget['jami']:
                        jami += budget['jami']
                        if budget2['ayollar']:
                            jami_ayol += budget2['ayollar']
                            jami_erkak += budget['jami'] - budget2['ayollar']
                        else:
                            jami_erkak += budget['jami']
                            jami_full_erkak += budget['jami']

                    if shartnoma['jami']:
                        jami += shartnoma['jami']
                        if budget3['ayollar']:
                            jami_ayol += budget3['ayollar']
                            jami_erkak += shartnoma['jami'] - budget3['ayollar']
                        else:
                            jami_erkak += shartnoma['jami']
                            jami_full_erkak += shartnoma['jami']
                    if k == 1:

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami_erkak)
                        ws.cell(row=row, column=9, value=jami_ayol)
                        kurs_1_jami += jami
                        kurs_1_erkak_jami += jami_erkak
                        kurs_1_ayol_jami += jami_ayol
                    elif k == 2:
                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami_erkak)
                        ws.cell(row=row, column=12, value=jami_ayol)
                        kurs_2_jami += jami
                        kurs_2_erkak_jami += jami_erkak
                        kurs_2_ayol_jami += jami_ayol
                    elif k == 3:

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami_erkak)
                        ws.cell(row=row, column=15, value=jami_ayol)
                        kurs_3_jami += jami
                        kurs_3_erkak_jami += jami_erkak
                        kurs_3_ayol_jami += jami_ayol
                    elif k == 4:

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami_erkak)
                        ws.cell(row=row, column=18, value=jami_ayol)
                        kurs_4_jami += jami
                        kurs_4_erkak_jami += jami_erkak
                        kurs_4_ayol_jami += jami_ayol
                    elif k == 5:

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami_erkak)
                        ws.cell(row=row, column=21, value=jami_ayol)
                        kurs_5_jami += jami
                        kurs_5_erkak_jami += jami_erkak
                        kurs_5_ayol_jami += jami_ayol
                    elif k == 6:

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami_erkak)
                        ws.cell(row=row, column=24, value=jami_ayol)
                        kurs_6_jami += jami
                        kurs_6_erkak_jami += jami_erkak
                        kurs_6_ayol_jami += jami_ayol
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=jami_erkak)
                        ws.cell(row=row, column=27, value=jami_ayol)
                        kurs_7_jami += jami
                        kurs_7_erkak_jami += jami_erkak
                        kurs_7_ayol_jami += jami_ayol
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)
            row += 1

            ws.cell(row=row, column=2, value='Sirtqi Jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_full_ayol)
            ws.cell(row=row, column=6, value=jami_full_erkak)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
            ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
            ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
            ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
            ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
            ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
            ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
            ws.cell(row=row, column=27, value=kurs_7_ayol_jami)

    row += 1
    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0
    for kunduzgi_yonlaish in yonalish_masofaviy:
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        budget2 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhbudjet__xotin_qiz'))
        budget3 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhshartnoma__xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']

                jami_erkak += budget['jami'] - budget2['ayollar']
                jami_full_ayol += budget2['ayollar']
                jami_full_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']

                jami_full_ayol += budget3['ayollar']
                jami_full_erkak += shartnoma['jami'] - budget3['ayollar']

            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)

            ws.cell(row=row, column=5, value=jami_ayol)
            ws.cell(row=row, column=6, value=jami_erkak)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    jami_erkak = 0
                    jami_ayol = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'), ayollar=Sum('xotin_qiz'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'),
                                                                                  ayollar=Sum('xotin_qiz'))
                    budget3 = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                       guruh__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    budget2 = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                    guruhi__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    if budget['jami']:
                        jami += budget['jami']
                        if budget2['ayollar']:
                            jami_ayol += budget2['ayollar']
                            jami_erkak += budget['jami'] - budget2['ayollar']
                        else:
                            jami_erkak += budget['jami']
                            jami_full_erkak += budget['jami']

                    if shartnoma['jami']:
                        jami += shartnoma['jami']
                        if budget3['ayollar']:
                            jami_ayol += budget3['ayollar']
                            jami_erkak += shartnoma['jami'] - budget3['ayollar']
                        else:
                            jami_erkak += shartnoma['jami']
                            jami_full_erkak += shartnoma['jami']
                    if k == 1:

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami_erkak)
                        ws.cell(row=row, column=9, value=jami_ayol)
                        kurs_1_jami += jami
                        kurs_1_erkak_jami += jami_erkak
                        kurs_1_ayol_jami += jami_ayol
                    elif k == 2:
                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami_erkak)
                        ws.cell(row=row, column=12, value=jami_ayol)
                        kurs_2_jami += jami
                        kurs_2_erkak_jami += jami_erkak
                        kurs_2_ayol_jami += jami_ayol
                    elif k == 3:

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami_erkak)
                        ws.cell(row=row, column=15, value=jami_ayol)
                        kurs_3_jami += jami
                        kurs_3_erkak_jami += jami_erkak
                        kurs_3_ayol_jami += jami_ayol
                    elif k == 4:

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami_erkak)
                        ws.cell(row=row, column=18, value=jami_ayol)
                        kurs_4_jami += jami
                        kurs_4_erkak_jami += jami_erkak
                        kurs_4_ayol_jami += jami_ayol
                    elif k == 5:

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami_erkak)
                        ws.cell(row=row, column=21, value=jami_ayol)
                        kurs_5_jami += jami
                        kurs_5_erkak_jami += jami_erkak
                        kurs_5_ayol_jami += jami_ayol
                    elif k == 6:

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami_erkak)
                        ws.cell(row=row, column=24, value=jami_ayol)
                        kurs_6_jami += jami
                        kurs_6_erkak_jami += jami_erkak
                        kurs_6_ayol_jami += jami_ayol
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=jami_erkak)
                        ws.cell(row=row, column=27, value=jami_ayol)
                        kurs_7_jami += jami
                        kurs_7_erkak_jami += jami_erkak
                        kurs_7_ayol_jami += jami_ayol
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)
            row += 1

            ws.cell(row=row, column=2, value='Masofaviy jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_full_ayol)
            ws.cell(row=row, column=6, value=jami_full_erkak)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
            ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
            ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
            ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
            ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
            ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
            ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
            ws.cell(row=row, column=27, value=kurs_7_ayol_jami)

    row += 1
    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0
    for kunduzgi_yonlaish in mut_2:
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        budget2 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhbudjet__xotin_qiz'))
        budget3 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhshartnoma__xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']

                jami_erkak += budget['jami'] - budget2['ayollar']
                jami_full_ayol += budget2['ayollar']
                jami_full_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']

                jami_full_ayol += budget3['ayollar']
                jami_full_erkak += shartnoma['jami'] - budget3['ayollar']

            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)

            ws.cell(row=row, column=5, value=jami_ayol)
            ws.cell(row=row, column=6, value=jami_erkak)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    jami_erkak = 0
                    jami_ayol = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'), ayollar=Sum('xotin_qiz'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'),
                                                                                  ayollar=Sum('xotin_qiz'))
                    budget3 = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                       guruh__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    budget2 = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                    guruhi__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    if budget['jami']:
                        jami += budget['jami']
                        if budget2['ayollar']:
                            jami_ayol += budget2['ayollar']
                            jami_erkak += budget['jami'] - budget2['ayollar']
                        else:
                            jami_erkak += budget['jami']
                            jami_full_erkak += budget['jami']

                    if shartnoma['jami']:
                        jami += shartnoma['jami']
                        if budget3['ayollar']:
                            jami_ayol += budget3['ayollar']
                            jami_erkak += shartnoma['jami'] - budget3['ayollar']
                        else:
                            jami_erkak += shartnoma['jami']
                            jami_full_erkak += shartnoma['jami']
                    if k == 1:

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami_erkak)
                        ws.cell(row=row, column=9, value=jami_ayol)
                        kurs_1_jami += jami
                        kurs_1_erkak_jami += jami_erkak
                        kurs_1_ayol_jami += jami_ayol
                    elif k == 2:
                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami_erkak)
                        ws.cell(row=row, column=12, value=jami_ayol)
                        kurs_2_jami += jami
                        kurs_2_erkak_jami += jami_erkak
                        kurs_2_ayol_jami += jami_ayol
                    elif k == 3:

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami_erkak)
                        ws.cell(row=row, column=15, value=jami_ayol)
                        kurs_3_jami += jami
                        kurs_3_erkak_jami += jami_erkak
                        kurs_3_ayol_jami += jami_ayol
                    elif k == 4:

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami_erkak)
                        ws.cell(row=row, column=18, value=jami_ayol)
                        kurs_4_jami += jami
                        kurs_4_erkak_jami += jami_erkak
                        kurs_4_ayol_jami += jami_ayol
                    elif k == 5:

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami_erkak)
                        ws.cell(row=row, column=21, value=jami_ayol)
                        kurs_5_jami += jami
                        kurs_5_erkak_jami += jami_erkak
                        kurs_5_ayol_jami += jami_ayol
                    elif k == 6:

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami_erkak)
                        ws.cell(row=row, column=24, value=jami_ayol)
                        kurs_6_jami += jami
                        kurs_6_erkak_jami += jami_erkak
                        kurs_6_ayol_jami += jami_ayol
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=jami_erkak)
                        ws.cell(row=row, column=27, value=jami_ayol)
                        kurs_7_jami += jami
                        kurs_7_erkak_jami += jami_erkak
                        kurs_7_ayol_jami += jami_ayol
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)
            row += 1

            ws.cell(row=row, column=2, value='2-mut jami')
            jami = 0
            jami_ayol = 0
            jami_erkak = 0
            kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs',
                                                                                             flat=True).distinct()
            budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'),
                                                                                         ayollar=Sum(
                                                                                             'guruhbudjet__xotin_qiz'))

            shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
                jami=Sum('guruhbudjet__jami'),
                ayollar=Sum(
                    'guruhbudjet__xotin_qiz'))
            if budget['jami']:
                jami += budget['jami']
                jami_full += budget['jami']
                if budget['ayollar']:
                    jami_ayol += budget['ayollar']

                    jami_erkak += budget['jami'] - budget['ayollar']
                    jami_full_ayol += shartnoma['ayollar']
                    jami_full_erkak += shartnoma['jami'] - shartnoma['ayollar']
                else:
                    jami_erkak += budget['jami']
                    jami_full_erkak += budget['jami']
            if shartnoma['jami']:
                jami += shartnoma['jami']
                jami_full += shartnoma['jami']
                if shartnoma['ayollar']:
                    jami_ayol += shartnoma['ayollar']
                    jami_erkak += shartnoma['jami'] - shartnoma['ayollar']

                    jami_full_ayol += shartnoma['ayollar']
                    jami_full_erkak += shartnoma['jami'] - shartnoma['ayollar']

                else:
                    jami_erkak += shartnoma['jami']
                    jami_full_erkak += shartnoma['jami']
            if jami != 0:
                ws.cell(row=row, column=1, value=row - 4)
                ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
                ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
                ws.cell(row=row, column=4, value=jami)

                ws.cell(row=row, column=5, value=jami_ayol)
                ws.cell(row=row, column=6, value=jami_erkak)

                for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                    if k in kurs_set:
                        jami = 0
                        jami_erkak = 0
                        jami_ayol = 0
                        budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                       guruhi__kurs=k).aggregate(jami=Sum('jami'),
                                                                                 ayollar=Sum('xotin_qiz'))
                        shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                             guruh__kurs=k).aggregate(jami=Sum('jami'),
                                                                                      ayollar=Sum('xotin_qiz'))
                        if budget['jami']:
                            jami += budget['jami']
                            if budget['ayollar']:
                                jami_ayol += budget['ayollar']
                                jami_erkak += budget['jami'] - budget['ayollar']
                            else:
                                jami_erkak += budget['jami']
                                jami_full_erkak += budget['jami']

                        if shartnoma['jami']:
                            jami += shartnoma['jami']
                            if shartnoma['ayollar']:
                                jami_ayol += shartnoma['ayollar']
                                jami_erkak += shartnoma['jami'] - shartnoma['ayollar']
                            else:
                                jami_erkak += shartnoma['jami']
                                jami_full_erkak += shartnoma['jami']

                        if k == 1:

                            ws.cell(row=row, column=7, value=jami)
                            ws.cell(row=row, column=8, value=jami_erkak)
                            ws.cell(row=row, column=9, value=jami_ayol)
                            kurs_1_jami += jami
                            kurs_1_erkak_jami += jami_erkak
                            kurs_1_ayol_jami += jami_ayol
                        elif k == 2:
                            ws.cell(row=row, column=10, value=jami)
                            ws.cell(row=row, column=11, value=jami_erkak)
                            ws.cell(row=row, column=12, value=jami_ayol)
                            kurs_2_jami += jami
                            kurs_2_erkak_jami += jami_erkak
                            kurs_2_ayol_jami += jami_ayol
                        elif k == 3:

                            ws.cell(row=row, column=13, value=jami)
                            ws.cell(row=row, column=14, value=jami_erkak)
                            ws.cell(row=row, column=15, value=jami_ayol)
                            kurs_3_jami += jami
                            kurs_3_erkak_jami += jami_erkak
                            kurs_3_ayol_jami += jami_ayol
                        elif k == 4:

                            ws.cell(row=row, column=16, value=jami)
                            ws.cell(row=row, column=17, value=jami_erkak)
                            ws.cell(row=row, column=18, value=jami_ayol)
                            kurs_4_jami += jami
                            kurs_4_erkak_jami += jami_erkak
                            kurs_4_ayol_jami += jami_ayol
                        elif k == 5:

                            ws.cell(row=row, column=19, value=jami)
                            ws.cell(row=row, column=20, value=jami_erkak)
                            ws.cell(row=row, column=21, value=jami_ayol)
                            kurs_5_jami += jami
                            kurs_5_erkak_jami += jami_erkak
                            kurs_5_ayol_jami += jami_ayol
                        elif k == 6:

                            ws.cell(row=row, column=22, value=jami)
                            ws.cell(row=row, column=23, value=jami_erkak)
                            ws.cell(row=row, column=24, value=jami_ayol)
                            kurs_6_jami += jami
                            kurs_6_erkak_jami += jami_erkak
                            kurs_6_ayol_jami += jami_ayol
                        else:
                            ws.cell(row=row, column=25, value=jami)
                            ws.cell(row=row, column=26, value=jami_erkak)
                            ws.cell(row=row, column=27, value=jami_ayol)
                            kurs_7_jami += jami
                            kurs_7_erkak_jami += jami_erkak
                            kurs_7_ayol_jami += jami_ayol
                    else:
                        # Handle missing course data
                        ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                        ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                        ws.cell(row=row, column=9 + (k - 1) * 3, value=0)
                row += 1

    row += 1
    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0
    for kunduzgi_yonlaish in magistir:
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        kurs_set = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).values_list('kurs', flat=True).distinct()
        budget = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(jami=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            jami=Sum('guruhshartnoma__jami'))
        budget2 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhbudjet__xotin_qiz'))
        budget3 = Guruh.objects.filter(org=org, yonalish=kunduzgi_yonlaish).aggregate(
            ayollar=Sum('guruhshartnoma__xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            jami_full += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']

                jami_erkak += budget['jami'] - budget2['ayollar']
                jami_full_ayol += budget2['ayollar']
                jami_full_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            jami_full += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']

                jami_full_ayol += budget3['ayollar']
                jami_full_erkak += shartnoma['jami'] - budget3['ayollar']

            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if jami != 0:
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=kunduzgi_yonlaish.name)
            ws.cell(row=row, column=3, value=kunduzgi_yonlaish.turi)
            ws.cell(row=row, column=4, value=jami)

            ws.cell(row=row, column=5, value=jami_ayol)
            ws.cell(row=row, column=6, value=jami_erkak)

            for k in [1, 2, 3, 4, 5, 6]:  # Iterate over predefined course values
                if k in kurs_set:
                    jami = 0
                    jami_erkak = 0
                    jami_ayol = 0
                    budget = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                   guruhi__kurs=k).aggregate(jami=Sum('jami'), ayollar=Sum('xotin_qiz'))
                    shartnoma = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                         guruh__kurs=k).aggregate(jami=Sum('jami'),
                                                                                  ayollar=Sum('xotin_qiz'))
                    budget3 = Shartnoma.objects.filter(guruh__org=org, guruh__yonalish=kunduzgi_yonlaish,
                                                       guruh__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    budget2 = Budjet.objects.filter(guruhi__org=org, guruhi__yonalish=kunduzgi_yonlaish,
                                                    guruhi__kurs=k).aggregate(ayollar=Sum('xotin_qiz'))
                    if budget['jami']:
                        jami += budget['jami']
                        if budget2['ayollar']:
                            jami_ayol += budget2['ayollar']
                            jami_erkak += budget['jami'] - budget2['ayollar']
                        else:
                            jami_erkak += budget['jami']
                            jami_full_erkak += budget['jami']

                    if shartnoma['jami']:
                        jami += shartnoma['jami']
                        if budget3['ayollar']:
                            jami_ayol += budget3['ayollar']
                            jami_erkak += shartnoma['jami'] - budget3['ayollar']
                        else:
                            jami_erkak += shartnoma['jami']
                            jami_full_erkak += shartnoma['jami']

                    if k == 1:

                        ws.cell(row=row, column=7, value=jami)
                        ws.cell(row=row, column=8, value=jami_erkak)
                        ws.cell(row=row, column=9, value=jami_ayol)
                        kurs_1_jami += jami
                        kurs_1_erkak_jami += jami_erkak
                        kurs_1_ayol_jami += jami_ayol
                    elif k == 2:
                        ws.cell(row=row, column=10, value=jami)
                        ws.cell(row=row, column=11, value=jami_erkak)
                        ws.cell(row=row, column=12, value=jami_ayol)
                        kurs_2_jami += jami
                        kurs_2_erkak_jami += jami_erkak
                        kurs_2_ayol_jami += jami_ayol
                    elif k == 3:

                        ws.cell(row=row, column=13, value=jami)
                        ws.cell(row=row, column=14, value=jami_erkak)
                        ws.cell(row=row, column=15, value=jami_ayol)
                        kurs_3_jami += jami
                        kurs_3_erkak_jami += jami_erkak
                        kurs_3_ayol_jami += jami_ayol
                    elif k == 4:

                        ws.cell(row=row, column=16, value=jami)
                        ws.cell(row=row, column=17, value=jami_erkak)
                        ws.cell(row=row, column=18, value=jami_ayol)
                        kurs_4_jami += jami
                        kurs_4_erkak_jami += jami_erkak
                        kurs_4_ayol_jami += jami_ayol
                    elif k == 5:

                        ws.cell(row=row, column=19, value=jami)
                        ws.cell(row=row, column=20, value=jami_erkak)
                        ws.cell(row=row, column=21, value=jami_ayol)
                        kurs_5_jami += jami
                        kurs_5_erkak_jami += jami_erkak
                        kurs_5_ayol_jami += jami_ayol
                    elif k == 6:

                        ws.cell(row=row, column=22, value=jami)
                        ws.cell(row=row, column=23, value=jami_erkak)
                        ws.cell(row=row, column=24, value=jami_ayol)
                        kurs_6_jami += jami
                        kurs_6_erkak_jami += jami_erkak
                        kurs_6_ayol_jami += jami_ayol
                    else:
                        ws.cell(row=row, column=25, value=jami)
                        ws.cell(row=row, column=26, value=jami_erkak)
                        ws.cell(row=row, column=27, value=jami_ayol)
                        kurs_7_jami += jami
                        kurs_7_erkak_jami += jami_erkak
                        kurs_7_ayol_jami += jami_ayol
                else:
                    # Handle missing course data
                    ws.cell(row=row, column=7 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=8 + (k - 1) * 3, value=0)
                    ws.cell(row=row, column=9 + (k - 1) * 3, value=0)
            row += 1

            ws.cell(row=row, column=2, value='Magistir jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_full_ayol)
            ws.cell(row=row, column=6, value=jami_full_erkak)
            ws.cell(row=row, column=7, value=kurs_1_jami)
            ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
            ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
            ws.cell(row=row, column=10, value=kurs_2_jami)
            ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
            ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
            ws.cell(row=row, column=13, value=kurs_3_jami)
            ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
            ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
            ws.cell(row=row, column=16, value=kurs_4_jami)
            ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
            ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
            ws.cell(row=row, column=19, value=kurs_5_jami)
            ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
            ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
            ws.cell(row=row, column=22, value=kurs_6_jami)
            ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
            ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
            ws.cell(row=row, column=25, value=kurs_7_jami)
            ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
            ws.cell(row=row, column=27, value=kurs_7_ayol_jami)
    row += 1
    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0

    guruhs = Guruh.objects.filter(org=org, bosqich='Bakalavr')
    budjet = Guruh.objects.filter(org=org, bosqich='Bakalavr').aggregate(budget=Sum('guruhbudjet__jami'))
    shartnoma = Guruh.objects.filter(org=org, bosqich='Bakalavr').aggregate(budget=Sum('guruhshartnoma__jami'))
    budget2 = Guruh.objects.filter(org=org, bosqich='Bakalavr').aggregate(ayollar=Sum('guruhbudjet__xotin_qiz'))
    budget3 = Guruh.objects.filter(org=org, bosqich='Bakalavr').aggregate(ayollar=Sum('guruhshartnoma__xotin_qiz'))
    jami = 0
    jami_ayol = 0
    jami_erkak = 0
    if budjet['budget']:
        jami += budjet['budget']
        jami_full += budjet['budget']
        if budget2['ayollar']:
            jami_ayol += budget2['ayollar']
            jami_erkak += budjet['budget'] - budget2['ayollar']
            jami_full_ayol += budget2['ayollar']
            jami_full_erkak += budjet['budget'] - budget2['ayollar']
        else:
            jami_erkak += budjet['budget']
            jami_full_erkak += budjet['budget']
    if shartnoma['budget']:
        jami += shartnoma['budget']
        jami_full += shartnoma['budget']
        if budget3['ayollar']:
            jami_ayol += budget3['ayollar']
            jami_erkak += shartnoma['budget'] - budget3['ayollar']
            jami_full_ayol += budget3['ayollar']
            jami_full_erkak += shartnoma['budget'] - budget3['ayollar']
        else:
            jami_erkak += shartnoma['budget']
            jami_full_erkak += shartnoma['budget']

    for guruh in guruhs:
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
        shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
        budget2 = Budjet.objects.filter(guruhi=guruh).aggregate(ayollar=Sum('xotin_qiz'))
        budget3 = Shartnoma.objects.filter(guruh=guruh).aggregate(ayollar=Sum('xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']
                jami_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']
            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if guruh.kurs == 1:
            kurs_1_jami += jami
            kurs_1_erkak_jami += jami_erkak
            kurs_1_ayol_jami += jami_ayol
        elif guruh.kurs == 2:
            kurs_2_jami += jami
            kurs_2_erkak_jami += jami_erkak
            kurs_2_ayol_jami += jami_ayol
        elif guruh.kurs == 3:
            kurs_3_jami += jami
            kurs_3_erkak_jami += jami_erkak
            kurs_3_ayol_jami += jami_ayol
        elif guruh.kurs == 4:
            kurs_4_jami += jami
            kurs_4_erkak_jami += jami_erkak
            kurs_4_ayol_jami += jami_ayol
        elif guruh.kurs == 5:
            kurs_5_jami += jami
            kurs_5_erkak_jami += jami_erkak
            kurs_5_ayol_jami += jami_ayol
        elif guruh.kurs == 6:
            kurs_6_jami += jami
            kurs_6_erkak_jami += jami_erkak
            kurs_6_ayol_jami += jami_ayol
        else:
            kurs_7_jami += jami
            kurs_7_erkak_jami += jami_erkak
            kurs_7_ayol_jami += jami_ayol

    row += 1
    ws.cell(row=row, column=2, value='Bakalavr jami')
    ws.cell(row=row, column=4, value=jami_full)
    ws.cell(row=row, column=5, value=jami_full_ayol)
    ws.cell(row=row, column=6, value=jami_full_erkak)
    ws.cell(row=row, column=7, value=kurs_1_jami)
    ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
    ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
    ws.cell(row=row, column=10, value=kurs_2_jami)
    ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
    ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
    ws.cell(row=row, column=13, value=kurs_3_jami)
    ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
    ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
    ws.cell(row=row, column=16, value=kurs_4_jami)
    ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
    ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
    ws.cell(row=row, column=19, value=kurs_5_jami)
    ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
    ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
    ws.cell(row=row, column=22, value=kurs_6_jami)
    ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
    ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
    ws.cell(row=row, column=25, value=kurs_7_jami)
    ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
    ws.cell(row=row, column=27, value=kurs_7_ayol_jami)


    row += 1

    fakultetlar = Faculty.objects.filter(org=org)
    for fakultet in fakultetlar:
        jami_full = 0
        kurs_1_jami = 0
        kurs_1_erkak_jami = 0
        kurs_1_ayol_jami = 0
        kurs_2_jami = 0
        kurs_2_erkak_jami = 0
        kurs_2_ayol_jami = 0
        kurs_3_jami = 0
        kurs_3_erkak_jami = 0
        kurs_3_ayol_jami = 0
        kurs_4_jami = 0
        kurs_4_erkak_jami = 0
        kurs_4_ayol_jami = 0
        kurs_5_jami = 0
        kurs_5_erkak_jami = 0
        kurs_5_ayol_jami = 0
        kurs_6_jami = 0
        kurs_6_erkak_jami = 0
        kurs_6_ayol_jami = 0
        kurs_7_jami = 0
        kurs_7_erkak_jami = 0
        kurs_7_ayol_jami = 0
        jami_full_erkak = 0
        jami_full_ayol = 0
        guruhs = Guruh.objects.filter(yonalish__faculty=fakultet)

        budjet = Guruh.objects.filter(yonalish__faculty=fakultet).aggregate(budget=Sum('guruhbudjet__jami'))
        shartnoma = Guruh.objects.filter(yonalish__faculty=fakultet).aggregate(budget=Sum('guruhshartnoma__jami'))
        budget2 = Guruh.objects.filter(yonalish__faculty=fakultet).aggregate(ayollar=Sum('guruhbudjet__xotin_qiz'))
        budget3 = Guruh.objects.filter(yonalish__faculty=fakultet).aggregate(ayollar=Sum('guruhshartnoma__xotin_qiz'))
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        if budjet['budget']:
            jami += budjet['budget']
            jami_full += budjet['budget']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']
                jami_erkak += budjet['budget'] - budget2['ayollar']
                jami_full_ayol += budget2['ayollar']
                jami_full_erkak += budjet['budget'] - budget2['ayollar']
            else:
                jami_erkak += budjet['budget']
                jami_full_erkak += budjet['budget']
        if shartnoma['budget']:
            jami += shartnoma['budget']
            jami_full += shartnoma['budget']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['budget'] - budget3['ayollar']
                jami_full_ayol += budget3['ayollar']
                jami_full_erkak += shartnoma['budget'] - budget3['ayollar']
            else:
                jami_erkak += shartnoma['budget']
                jami_full_erkak += shartnoma['budget']
        for guruh in guruhs:
            jami = 0
            jami_ayol = 0
            jami_erkak = 0
            budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
            shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
            budget2 = Budjet.objects.filter(guruhi=guruh).aggregate(ayollar=Sum('xotin_qiz'))
            budget3 = Shartnoma.objects.filter(guruh=guruh).aggregate(ayollar=Sum('xotin_qiz'))
            if budget['jami']:
                jami += budget['jami']
                if budget2['ayollar']:
                    jami_ayol += budget2['ayollar']
                    jami_erkak += budget['jami'] - budget2['ayollar']
                else:
                    jami_erkak += budget['jami']
                    jami_full_erkak += budget['jami']
            if shartnoma['jami']:
                jami += shartnoma['jami']
                if budget3['ayollar']:
                    jami_ayol += budget3['ayollar']
                    jami_erkak += shartnoma['jami'] - budget3['ayollar']
                else:
                    jami_erkak += shartnoma['jami']
                    jami_full_erkak += shartnoma['jami']

            if guruh.kurs == 1:
                kurs_1_jami += jami
                kurs_1_erkak_jami += jami_erkak
                kurs_1_ayol_jami += jami_ayol
            elif guruh.kurs == 2:
                kurs_2_jami += jami
                kurs_2_erkak_jami += jami_erkak
                kurs_2_ayol_jami += jami_ayol
            elif guruh.kurs == 3:
                kurs_3_jami += jami
                kurs_3_erkak_jami += jami_erkak
                kurs_3_ayol_jami += jami_ayol
            elif guruh.kurs == 4:
                kurs_4_jami += jami
                kurs_4_erkak_jami += jami_erkak
                kurs_4_ayol_jami += jami_ayol
            elif guruh.kurs == 5:
                kurs_5_jami += jami
                kurs_5_erkak_jami += jami_erkak
                kurs_5_ayol_jami += jami_ayol
            elif guruh.kurs == 6:
                kurs_6_jami += jami
                kurs_6_erkak_jami += jami_erkak
                kurs_6_ayol_jami += jami_ayol
            else:
                kurs_7_jami += jami
                kurs_7_erkak_jami += jami_erkak
                kurs_7_ayol_jami += jami_ayol


        ws.cell(row=row, column=2, value=f'{fakultet.name}')
        ws.cell(row=row, column=4, value=jami_full)
        ws.cell(row=row, column=5, value=jami_full_ayol)
        ws.cell(row=row, column=6, value=jami_full_erkak)
        ws.cell(row=row, column=7, value=kurs_1_jami)
        ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
        ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
        ws.cell(row=row, column=10, value=kurs_2_jami)
        ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
        ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
        ws.cell(row=row, column=13, value=kurs_3_jami)
        ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
        ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
        ws.cell(row=row, column=16, value=kurs_4_jami)
        ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
        ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
        ws.cell(row=row, column=19, value=kurs_5_jami)
        ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
        ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
        ws.cell(row=row, column=22, value=kurs_6_jami)
        ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
        ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
        ws.cell(row=row, column=25, value=kurs_7_jami)
        ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
        ws.cell(row=row, column=27, value=kurs_7_ayol_jami)
        row += 1

    row += 1
    jami_full = 0
    kurs_1_jami = 0
    kurs_1_erkak_jami = 0
    kurs_1_ayol_jami = 0
    kurs_2_jami = 0
    kurs_2_erkak_jami = 0
    kurs_2_ayol_jami = 0
    kurs_3_jami = 0
    kurs_3_erkak_jami = 0
    kurs_3_ayol_jami = 0
    kurs_4_jami = 0
    kurs_4_erkak_jami = 0
    kurs_4_ayol_jami = 0
    kurs_5_jami = 0
    kurs_5_erkak_jami = 0
    kurs_5_ayol_jami = 0
    kurs_6_jami = 0
    kurs_6_erkak_jami = 0
    kurs_6_ayol_jami = 0
    kurs_7_jami = 0
    kurs_7_erkak_jami = 0
    kurs_7_ayol_jami = 0
    jami_full_erkak = 0
    jami_full_ayol = 0
    guruhs = Guruh.objects.filter(org=org)

    budjet = Guruh.objects.filter(org=org).aggregate(budget=Sum('guruhbudjet__jami'))
    shartnoma = Guruh.objects.filter(org=org).aggregate(budget=Sum('guruhshartnoma__jami'))
    budget2 = Guruh.objects.filter(org=org).aggregate(ayollar=Sum('guruhbudjet__xotin_qiz'))
    budget3 = Guruh.objects.filter(org=org).aggregate(ayollar=Sum('guruhshartnoma__xotin_qiz'))
    jami = 0
    jami_ayol = 0
    jami_erkak = 0
    if budjet['budget']:
        jami += budjet['budget']
        jami_full += budjet['budget']
        if budget2['ayollar']:
            jami_ayol += budget2['ayollar']
            jami_erkak += budjet['budget'] - budget2['ayollar']
            jami_full_ayol += budget2['ayollar']
            jami_full_erkak += budjet['budget'] - budget2['ayollar']
        else:
            jami_erkak += budjet['budget']
            jami_full_erkak += budjet['budget']
    if shartnoma['budget']:
        jami += shartnoma['budget']
        jami_full += shartnoma['budget']
        if budget3['ayollar']:
            jami_ayol += budget3['ayollar']
            jami_erkak += shartnoma['budget'] - budget3['ayollar']
            jami_full_ayol += budget3['ayollar']
            jami_full_erkak += shartnoma['budget'] - budget3['ayollar']
        else:
            jami_erkak += shartnoma['budget']
            jami_full_erkak += shartnoma['budget']
    for guruh in guruhs:
        jami = 0
        jami_ayol = 0
        jami_erkak = 0
        budget = Budjet.objects.filter(guruhi=guruh).aggregate(jami=Sum('jami'))
        shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(jami=Sum('jami'))
        budget2 = Budjet.objects.filter(guruhi=guruh).aggregate(ayollar=Sum('xotin_qiz'))
        budget3 = Shartnoma.objects.filter(guruh=guruh).aggregate(ayollar=Sum('xotin_qiz'))
        if budget['jami']:
            jami += budget['jami']
            if budget2['ayollar']:
                jami_ayol += budget2['ayollar']
                jami_erkak += budget['jami'] - budget2['ayollar']
            else:
                jami_erkak += budget['jami']
                jami_full_erkak += budget['jami']
        if shartnoma['jami']:
            jami += shartnoma['jami']
            if budget3['ayollar']:
                jami_ayol += budget3['ayollar']
                jami_erkak += shartnoma['jami'] - budget3['ayollar']
            else:
                jami_erkak += shartnoma['jami']
                jami_full_erkak += shartnoma['jami']

        if guruh.kurs == 1:
            kurs_1_jami += jami
            kurs_1_erkak_jami += jami_erkak
            kurs_1_ayol_jami += jami_ayol
        elif guruh.kurs == 2:
            kurs_2_jami += jami
            kurs_2_erkak_jami += jami_erkak
            kurs_2_ayol_jami += jami_ayol
        elif guruh.kurs == 3:
            kurs_3_jami += jami
            kurs_3_erkak_jami += jami_erkak
            kurs_3_ayol_jami += jami_ayol
        elif guruh.kurs == 4:
            kurs_4_jami += jami
            kurs_4_erkak_jami += jami_erkak
            kurs_4_ayol_jami += jami_ayol
        elif guruh.kurs == 5:
            kurs_5_jami += jami
            kurs_5_erkak_jami += jami_erkak
            kurs_5_ayol_jami += jami_ayol
        elif guruh.kurs == 6:
            kurs_6_jami += jami
            kurs_6_erkak_jami += jami_erkak
            kurs_6_ayol_jami += jami_ayol
        else:
            kurs_7_jami += jami
            kurs_7_erkak_jami += jami_erkak
            kurs_7_ayol_jami += jami_ayol

    ws.cell(row=row, column=2, value=f'Jami')
    ws.cell(row=row, column=4, value=jami_full)
    ws.cell(row=row, column=5, value=jami_full_ayol)
    ws.cell(row=row, column=6, value=jami_full_erkak)
    ws.cell(row=row, column=7, value=kurs_1_jami)
    ws.cell(row=row, column=8, value=kurs_1_erkak_jami)
    ws.cell(row=row, column=9, value=kurs_1_ayol_jami)
    ws.cell(row=row, column=10, value=kurs_2_jami)
    ws.cell(row=row, column=11, value=kurs_2_erkak_jami)
    ws.cell(row=row, column=12, value=kurs_2_ayol_jami)
    ws.cell(row=row, column=13, value=kurs_3_jami)
    ws.cell(row=row, column=14, value=kurs_3_erkak_jami)
    ws.cell(row=row, column=15, value=kurs_3_ayol_jami)
    ws.cell(row=row, column=16, value=kurs_4_jami)
    ws.cell(row=row, column=17, value=kurs_4_erkak_jami)
    ws.cell(row=row, column=18, value=kurs_4_ayol_jami)
    ws.cell(row=row, column=19, value=kurs_5_jami)
    ws.cell(row=row, column=20, value=kurs_5_erkak_jami)
    ws.cell(row=row, column=21, value=kurs_5_ayol_jami)
    ws.cell(row=row, column=22, value=kurs_6_jami)
    ws.cell(row=row, column=23, value=kurs_6_erkak_jami)
    ws.cell(row=row, column=24, value=kurs_6_ayol_jami)
    ws.cell(row=row, column=25, value=kurs_7_jami)
    ws.cell(row=row, column=26, value=kurs_7_erkak_jami)
    ws.cell(row=row, column=27, value=kurs_7_ayol_jami)
    row += 1

    wb.save(output4)
    output4.seek(0)
    return output4



