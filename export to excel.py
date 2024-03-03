from datetime import datetime

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import os
import pandas as pd
from openpyxl.utils import get_column_letter

FONT_NAME = 'Times New Roman'
FONT_SIZE = 12
FONT_COLOR = 'FF000000'
BORDER_STYLE = 'thin'
BORDER_COLOR = 'FF000000'


def set_cell_properties(cell, value, alignment, font, border):
    cell.value = value
    cell.alignment = alignment
    cell.font = font
    cell.border = border


wb = Workbook()
ws = wb.active
# add auto size width of column

# Hozirgi vaqtni olib, formatini belgilash


now = datetime.now()
formatted_time = now.strftime("%Y-%m-%d")
organization_name = "Toshkent Davlat Texnnika Universiteti"

ws.merge_cells('A1:P2')
set_cell_properties(ws.cell(row=1, column=1),
                    f"{organization_name} talabalari kontingentining {formatted_time} holati haqida umumiy ma'lumot",
                    Alignment(horizontal='center', vertical='center'),
                    Font(name=FONT_NAME, size=14, bold=True, italic=False, color=FONT_COLOR),
                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

cell_properties = [
    {'row': 3, 'column': 3, 'value': 'Guruhlar', 'width': True, 'merge': 'C3:C4'},
    {'row': 3, 'column': 4, 'value': 'Yo\'nalish', 'width': True, 'merge': 'D3:D4'},
    {'row': 3, 'column': 5, 'value': 'Jami', 'width': True, 'merge': 'E3:E4'},
    {'row': 3, 'column': 6, 'value': 'Shundan qizlar', 'width': True, 'merge': 'F3:F4'},
    {'row': 3, 'column': 7, 'value': 'Harbiy', 'width': True, 'merge': 'G3:G4'},
    {'row': 3, 'column': 8, 'value': 'Yangi qo\'shilgan', 'width': True, 'merge': 'H3:H4'},
    {'row': 3, 'column': 9, 'value': 'Chetlashtirilgan', 'width': True, 'merge': 'I3:I4'},
    {'row': 3, 'column': 10, 'value': 'Akademik tatil', 'width': True, 'merge': 'J3:J4'},
    {'row': 3, 'column': 11, 'value': 'Byudjet', 'width': True, 'merge': 'K3:M3'},
    {'row': 4, 'column': 11, 'value': 'Jami', 'width': True},
    {'row': 4, 'column': 12, 'value': 'Harbiy', 'width': True},
    {'row': 4, 'column': 13, 'value': 'Xotn qizlar', 'width': True},
    {'row': 3, 'column': 14, 'value': 'Shartmona', 'width': True, 'merge': 'N3:P3'},
    {'row': 4, 'column': 14, 'value': 'Jami', 'width': True},
    {'row': 4, 'column': 15, 'value': 'Harbiy', 'width': True},
    {'row': 4, 'column': 16, 'value': 'Xotn qizlar', 'width': True},

    # Add more dictionaries for other cells
]

for properties in cell_properties:
    cell = ws.cell(row=properties['row'], column=properties['column'])
    set_cell_properties(cell,
                        properties['value'],
                        Alignment(horizontal='center', vertical='center'),
                        Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
    if 'width' in properties and properties['width']:
        ws.column_dimensions[get_column_letter(properties['column'])].auto_size = True
    if 'merge' in properties:
        ws.merge_cells(properties['merge'])

coluns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
for col in coluns:
    ws.column_dimensions[col].width = 17

#
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'condig.settings')

django.setup()
from user.models import  Group, Faculty, Budjet, Shartnoma,Organization,Yonalish,Guruh

from django.contrib.auth.models import User

org = OrganizationModel.objects.filter(name='kiuf').first()


def find_kurs(org):
    return Kurs.objects.filter(organization=org).all().order_by('name')


# Kurslar bo'yicha ma'lumotlarni yozish

# filter by kurs and hullas C5:P5 da merge bo'lishi kerka va bu yerga kur nomi yoziwsh kerak va shu kursga tegishli bo'glgan guruhlarni filtir qilishi kerak va C6
# dan boshlab guruhlarni nomini yozish kerak shu tartibda shu gruhlarni filedlarni tartib bo'yicha yozilishi kerak. kegin bu kursni guruhlari tugagandan keyin boshqa kursga o'tishi kerak va shu tartibda yozilishi kerak
#
# def write_kurs_data(org, ws):
#     kurslar = find_kurs(org)
#     row = 5
#     for kurs in kurslar:
#         ws.merge_cells(f'C{row}:P{row}')
#         set_cell_properties(ws.cell(row=row, column=3),
#                             kurs.name,
#                             Alignment(horizontal='center', vertical='center'),
#                             Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
#                             Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                    right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                    top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                    bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#
#         groups = Group.objects.filter(organization=org, kurs=kurs).all()
#         row += 1  # Shu yerda qatorni o'tkazamiz, keyinroq kelgan kurslarga joy bermoqchi bo'lsangiz
#         for group in groups:
#             set_cell_properties(ws.cell(row=row, column=3),
#                                 group.name,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             row += 1
#
#
# def write_kurs_data(org, ws):
#     kurslar = find_kurs(org)
#     row = 5
#     for kurs in kurslar:
#         ws.merge_cells(f'C{row}:P{row}')
#         set_cell_properties(ws.cell(row=row, column=3),
#                             kurs.name,
#                             Alignment(horizontal='center', vertical='center'),
#                             Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
#                             Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                    right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                    top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                    bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#
#         faculty = Faculty.objects.filter(organization=org).first()
#
#         groups = Group.objects.filter(organization=org, kurs=kurs).all()
#
#         row += 1  # Shu yerda qatorni o'tkazamiz, keyinroq kelgan kurslarga joy bermoqchi bo'lsangiz
#         for group in groups:
#             set_cell_properties(ws.cell(row=row, column=3),
#                                 group.name,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#
#             set_cell_properties(ws.cell(row=row, column=4),
#                                 group.faculty.name,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Jami budjetni jami + shartnoma jami qo'shiladi
#             # ws.cell(row=row, column=5, value=group.jami)
#             budjet_jami = group.budjetgroup.aggregate(Sum('jami')).get('jami__sum') or 0
#             shartnoma_jami = group.shartnomagroup.aggregate(Sum('jami')).get('jami__sum') or 0
#
#             set_cell_properties(ws.cell(row=row, column=5),
#                                 budjet_jami + shartnoma_jami,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             set_cell_properties(ws.cell(row=row, column=5),
#                                 budjet_jami + shartnoma_jami,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Shundan qizlar
#             xotn_qizlar = group.budjetgroup.aggregate(Sum('xotn_qizlar')).get('xotn_qizlar__sum') or 0
#             xotn_qizlar += group.shartnomagroup.aggregate(Sum('xotn_qizlar')).get('xotn_qizlar__sum') or 0
#             set_cell_properties(ws.cell(row=row, column=6),
#                                 xotn_qizlar,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Harbiy
#             harbiy = group.budjetgroup.aggregate(Sum('harbiy')).get('harbiy__sum') or 0
#             harbiy += group.shartnomagroup.aggregate(Sum('harbiy')).get('harbiy__sum') or 0
#             set_cell_properties(ws.cell(row=row, column=7),
#                                 harbiy,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Yangi qo'shilgan
#             set_cell_properties(ws.cell(row=row, column=8),
#                                 group.newstudnets,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Chetlashtirilgan
#             set_cell_properties(ws.cell(row=row, column=9),
#                                 group.chetlashtirilgan,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Akademik tatil
#             set_cell_properties(ws.cell(row=row, column=10),
#                                 group.akademik,
#                                 Alignment(horizontal='center', vertical='center'),
#                                 Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                 Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                        bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Byudjet
#             budjet = Budjet.objects.filter(group=group).first()
#             if budjet:
#                 # Jami
#                 set_cell_properties(ws.cell(row=row, column=11),
#                                     budjet.jami,
#                                     Alignment(horizontal='center', vertical='center'),
#                                     Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#                 # Harbiy
#                 set_cell_properties(ws.cell(row=row, column=12),
#                                     budjet.harbiy,
#                                     Alignment(horizontal='center', vertical='center'),
#                                     Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#                 # Xotn qizlar
#                 set_cell_properties(ws.cell(row=row, column=13),
#                                     budjet.xotn_qizlar,
#                                     Alignment(horizontal='center', vertical='center'),
#                                     Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#             # Shartmona
#             shartmona = Shartmonoma.objects.filter(group=group).first()
#             if shartmona:
#                 # Jami
#                 set_cell_properties(ws.cell(row=row, column=14),
#                                     shartmona.jami,
#                                     Alignment(horizontal='center', vertical='center'),
#                                     Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#                 # Harbiy
#                 set_cell_properties(ws.cell(row=row, column=15),
#                                     shartmona.harbiy,
#                                     Alignment(horizontal='center', vertical='center'),
#                                     Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#                 # Xotn qizlar
#                 set_cell_properties(ws.cell(row=row, column=16),
#                                     shartmona.xotn_qizlar,
#                                     Alignment(horizontal='center', vertical='center'),
#                                     Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
#                                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
#                                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
#
#             row += 1
def write_kurs_data(org, ws):
    kurslar = find_kurs(org)
    row = 5

    for kurs in kurslar:
        ws.merge_cells(f'C{row}:P{row}')
        set_cell_properties(ws.cell(row=row, column=3),
                            kurs.name,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

        faculty = Faculty.objects.filter(organization=org).first()

        groups = Group.objects.filter(organization=org, kurs=kurs).all()

        row += 1  # Shu yerda qatorni o'tkazamiz, keyinroq kelgan kurslarga joy bermoqchi bo'lsangiz

        kurs_jami = 0
        kurs_xotn_qizlar = 0
        kurs_harbiy = 0
        kurs_newstudnets = 0
        kurs_chetlashtirilgan = 0
        kurs_akademik = 0
        kurs_budjet_jami = 0
        kurs_budjet_harbiy = 0
        kurs_budjet_xotn_qizlar = 0
        kurs_shartnoma_jami = 0
        kurs_shartnoma_harbiy = 0
        kurs_shartnoma_xotn_qizlar = 0

        for group in groups:
            set_cell_properties(ws.cell(row=row, column=3),
                                group.name,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            set_cell_properties(ws.cell(row=row, column=4),
                                group.faculty.name,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Jami budjetni jami + shartnoma jami qo'shiladi
            budjet_jami = group.budjetgroup.aggregate(Sum('jami')).get('jami__sum') or 0
            shartnoma_jami = group.shartnomagroup.aggregate(Sum('jami')).get('jami__sum') or 0

            group_jami = budjet_jami + shartnoma_jami
            kurs_jami += group_jami

            set_cell_properties(ws.cell(row=row, column=5),
                                group_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Shundan qizlar
            xotn_qizlar = group.budjetgroup.aggregate(Sum('xotn_qizlar')).get('xotn_qizlar__sum') or 0
            xotn_qizlar += group.shartnomagroup.aggregate(Sum('xotn_qizlar')).get('xotn_qizlar__sum') or 0
            kurs_xotn_qizlar += xotn_qizlar

            set_cell_properties(ws.cell(row=row, column=6),
                                xotn_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Harbiy
            harbiy = group.budjetgroup.aggregate(Sum('harbiy')).get('harbiy__sum') or 0
            harbiy += group.shartnomagroup.aggregate(Sum('harbiy')).get('harbiy__sum') or 0
            kurs_harbiy += harbiy

            set_cell_properties(ws.cell(row=row, column=7),
                                harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Yangi qo'shilgan
            kurs_newstudnets += group.newstudnets

            set_cell_properties(ws.cell(row=row, column=8),
                                group.newstudnets,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Chetlashtirilgan
            kurs_chetlashtirilgan += group.chetlashtirilgan

            set_cell_properties(ws.cell(row=row, column=9),
                                group.chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Akademik tatil
            kurs_akademik += group.akademik

            set_cell_properties(ws.cell(row=row, column=10),
                                group.akademik,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Byudjet
            budjet = Budjet.objects.filter(group=group).first()
            if budjet:
                # Jami
                kurs_budjet_jami += budjet.jami

                set_cell_properties(ws.cell(row=row, column=11),
                                    budjet.jami,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                # Harbiy
                kurs_budjet_harbiy += budjet.harbiy

                set_cell_properties(ws.cell(row=row, column=12),
                                    budjet.harbiy,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                # Xotn qizlar
                kurs_budjet_xotn_qizlar += budjet.xotn_qizlar

                set_cell_properties(ws.cell(row=row, column=13),
                                    budjet.xotn_qizlar,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            # Shartmona
            shartmona = Shartmonoma.objects.filter(group=group).first()
            if shartmona:
                # Jami
                kurs_shartnoma_jami += shartmona.jami

                set_cell_properties(ws.cell(row=row, column=14),
                                    shartmona.jami,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                # Harbiy
                kurs_shartnoma_harbiy += shartmona.harbiy

                set_cell_properties(ws.cell(row=row, column=15),
                                    shartmona.harbiy,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                # Xotn qizlar
                kurs_shartnoma_xotn_qizlar += shartmona.xotn_qizlar

                set_cell_properties(ws.cell(row=row, column=16),
                                    shartmona.xotn_qizlar,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

            row += 1

        # Kurs uchun jami qatorini qo'shamiz
        # shu yerda har bitta fakultet tugaganda shu kurs bo'yicha hamma filedlarni hisoblash kerak
        row += 1  # Yangi qator qo'shish uchun
        # set_cell_properties(ws.cell(row=row, column=3),
        #                     f"{kurs.name} bo'yicha jami",
        #                     Alignment(horizontal='center', vertical='center'),
        #                     Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
        #                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
        #                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
        #                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
        #                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
        #
        # set_cell_properties(ws.cell(row=row, column=5),
        #                     kurs_jami,
        #                     Alignment(horizontal='center', vertical='center'),
        #                     Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
        #                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
        #                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
        #                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
        #                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))


write_kurs_data(org, ws)

if 'talabalar.xlsx' in os.listdir():
    os.remove('talabalar.xlsx')

wb.save('talabalar.xlsx')
