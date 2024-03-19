import os

from io import BytesIO

import django
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()

from datetime import datetime

from django.db.models import Sum

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()
from user.models import Faculty, Budjet, Shartnoma, Organization, Yonalish, Guruh




def exporttoexcel6(org):
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl import Workbook

    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")

    def set_cell_properties(cell, value, alignment, font, border):
        cell.value = value
        cell.alignment = alignment
        cell.font = font
        cell.border = border

    FONT_NAME = 'Times New Roman'
    FONT_SIZE = 8
    FONT_COLOR = 'FF000000'
    BORDER_STYLE = 'thin'
    BORDER_COLOR = 'FF000000'
    red_color = 'FF0000FF'
    # color blue
    blue_color = 'FFFF0000'

    # add auto size width of column

    # Hozirgi vaqtni ol
    # add auto size width of column

    # Hozirgi vaqtni olib, formatini belgilash

    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")
    wb = Workbook()
    ws = wb.active
    output4 = BytesIO()
    organization_name = org.full_name

    ws.merge_cells('A1:D2')
    set_cell_properties(ws.cell(row=1, column=1),
                        f"{organization_name} talabalari kontingentining {formatted_time} holati haqida umumiy ma'lumot ",
                        Alignment(horizontal='center', vertical='center'),
                        Font(name=FONT_NAME, size=8, bold=True, italic=False, color=FONT_COLOR),
                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

    cell_properties = [
        {'row': 3, 'column': 1, 'value': 'Kurslar kesimida talabalar soni', 'width': True, 'merge': 'A3:D3'},
        {'row': 4, 'column': 1, 'value': 'Kurs', 'width': True},
        {'row': 4, 'column': 2, 'value': 'Talaba soni', 'width': True},
        {'row': 4, 'column': 3, 'value': 'O\'g\'il', 'width': True},
        {'row': 4, 'column': 4, 'value': 'Qiz', 'width': True},

    ]

    for properties in cell_properties:
        cell = ws.cell(row=properties['row'], column=properties['column'])
        set_cell_properties(cell,
                            properties['value'],
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
        if 'width' in properties and properties['width']:
            ws.column_dimensions[get_column_letter(properties['column'])].auto_size = True
        if 'merge' in properties:
            ws.merge_cells(properties['merge'])
    wb.save('talabalar.xlsx')
    output4.seek(0)

    row = 5
    kurs_set = Guruh.objects.filter(org=org, bosqich='Bakalavr', yonalish__turi='Kunduzgi').exclude(
        yonalish__mutahasislik_2=True).values_list(
        'kurs', flat=True).distinct()
    bakalavr_guruhs = Guruh.objects.filter(org=org, bosqich='Bakalavr', yonalish__turi='Kunduzgi').exclude(
        yonalish__mutahasislik_2=True)
    jami = 0
    erkak = 0
    ayol = 0
    ws.cell(row=row, column=1, value='Talaba soni (Bakalavriat)').alignment = Alignment(horizontal='center',
                                                                                        vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    # shakl background rangini och havo rang qil
    och_havo = PatternFill(start_color="b3dfe8", end_color="b3dfe8", fill_type="solid")
    ws.cell(row=row, column=1).fill = och_havo
    row += 1

    for k in [1, 2, 3, 4, 5, 6, 7]:
        jami_guruh = 0
        erkak_jami = 0
        ayol_jami = 0
        for guruh in bakalavr_guruhs:
            if k in kurs_set:
                budget = Budjet.objects.filter(guruhi=guruh, guruhi__kurs=k).aggregate(Sum('jami'), Sum('xotin_qiz'))
                shartnoma = Shartnoma.objects.filter(guruh=guruh, guruh__kurs=k).aggregate(Sum('jami'),
                                                                                           Sum('xotin_qiz'))
                if budget['jami__sum'] is not None:
                    jami_guruh += budget['jami__sum']
                    erkak_jami += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol_jami += budget['xotin_qiz__sum']
                    jami += budget['jami__sum']
                    erkak += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol += budget['xotin_qiz__sum']
                if shartnoma['jami__sum'] is not None:
                    jami_guruh += shartnoma['jami__sum']
                    erkak_jami += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol_jami += shartnoma['xotin_qiz__sum']
                    jami += shartnoma['jami__sum']
                    erkak += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol += shartnoma['xotin_qiz__sum']
        if jami_guruh != 0:
            ws.cell(row=row, column=1, value=k).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2, value=jami_guruh).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value=erkak_jami).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=ayol_jami).alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    ws.cell(row=row, column=1, value='Bakalavr jami').alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=jami).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=erkak).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=ayol).alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    kurs_set = Guruh.objects.filter(org=org, yonalish__turi='Sirtqi').exclude(yonalish__mutahasislik_2=True,
                                                                              yonalish__turi='Masofaviy',
                                                                              bosqich='Magistr').values_list(
        'kurs', flat=True).distinct()
    bakalavr_guruhs = Guruh.objects.filter(org=org, yonalish__turi='Sirtqi').exclude(yonalish__mutahasislik_2=True,
                                                                                     yonalish__turi='Masofaviy',
                                                                                     bosqich='Magistr').distinct()
    bakalavr_guruhs = bakalavr_guruhs.exclude(yonalish__turi='Kunduzgi')
    jami = 0
    erkak = 0
    ayol = 0
    ws.cell(row=row, column=1, value='Talaba soni (Sirtqi)').alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    # shakl background rangini och havo rang qil
    och_havo = PatternFill(start_color="b3dfe8", end_color="b3dfe8", fill_type="solid")
    ws.cell(row=row, column=1).fill = och_havo
    row += 1

    for k in [1, 2, 3, 4, 5, 6, 7]:
        jami_guruh = 0
        erkak_jami = 0
        ayol_jami = 0
        for guruh in bakalavr_guruhs:
            if k in kurs_set:
                budget = Budjet.objects.filter(guruhi=guruh, guruhi__kurs=k).aggregate(Sum('jami'), Sum('xotin_qiz'))
                shartnoma = Shartnoma.objects.filter(guruh=guruh, guruh__kurs=k).aggregate(Sum('jami'),
                                                                                           Sum('xotin_qiz'))
                if budget['jami__sum'] is not None:
                    jami_guruh += budget['jami__sum']
                    erkak_jami += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol_jami += budget['xotin_qiz__sum']
                    jami += budget['jami__sum']
                    erkak += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol += budget['xotin_qiz__sum']
                if shartnoma['jami__sum'] is not None:
                    jami_guruh += shartnoma['jami__sum']
                    erkak_jami += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol_jami += shartnoma['xotin_qiz__sum']
                    jami += shartnoma['jami__sum']
                    erkak += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol += shartnoma['xotin_qiz__sum']
        if jami_guruh != 0:
            ws.cell(row=row, column=1, value=k).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2, value=jami_guruh).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value=erkak_jami).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=ayol_jami).alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    ws.cell(row=row, column=1, value='Sirtqi jami').alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=jami).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=erkak).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=ayol).alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    kurs_set = Guruh.objects.filter(org=org, yonalish__mutahasislik_2=True).exclude(bosqich='Magistr',
                                                                                    yonalish__turi='Masofaviy').values_list(
        'kurs', flat=True).distinct()
    bakalavr_guruhs = Guruh.objects.filter(org=org, yonalish__mutahasislik_2=True).exclude(bosqich='Magistr',
                                                                                           yonalish__turi='Masofaviy').distinct()
    jami = 0
    erkak = 0
    ayol = 0
    ws.cell(row=row, column=1, value='Talaba soni (2-mutaxasislik)').alignment = Alignment(horizontal='center',
                                                                                           vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    # shakl background rangini och havo rang qil
    och_havo = PatternFill(start_color="b3dfe8", end_color="b3dfe8", fill_type="solid")
    ws.cell(row=row, column=1).fill = och_havo
    row += 1

    for k in [1, 2, 3, 4, 5, 6, 7]:
        jami_guruh = 0
        erkak_jami = 0
        ayol_jami = 0
        for guruh in bakalavr_guruhs:
            if k in kurs_set:
                budget = Budjet.objects.filter(guruhi=guruh, guruhi__kurs=k).aggregate(Sum('jami'), Sum('xotin_qiz'))
                shartnoma = Shartnoma.objects.filter(guruh=guruh, guruh__kurs=k).aggregate(Sum('jami'),
                                                                                           Sum('xotin_qiz'))
                if budget['jami__sum'] is not None:
                    jami_guruh += budget['jami__sum']
                    erkak_jami += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol_jami += budget['xotin_qiz__sum']
                    jami += budget['jami__sum']
                    erkak += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol += budget['xotin_qiz__sum']
                if shartnoma['jami__sum'] is not None:
                    jami_guruh += shartnoma['jami__sum']
                    erkak_jami += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol_jami += shartnoma['xotin_qiz__sum']
                    jami += shartnoma['jami__sum']
                    erkak += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol += shartnoma['xotin_qiz__sum']
        if jami_guruh != 0:
            ws.cell(row=row, column=1, value=k).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2, value=jami_guruh).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value=erkak_jami).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=ayol_jami).alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    ws.cell(row=row, column=1, value='2-mut jami').alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=jami).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=erkak).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=ayol).alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    kurs_set = Guruh.objects.filter(org=org, yonalish__turi='Masofaviy').exclude(yonalish__mutahasislik_2=True,
                                                                                 bosqich='Magistr').values_list(
        'kurs', flat=True).distinct()
    bakalavr_guruhs = Guruh.objects.filter(org=org, yonalish__turi='Masofaviy').exclude(yonalish__mutahasislik_2=True,
                                                                                        bosqich='Magistr').distinct()
    jami = 0
    erkak = 0
    ayol = 0
    ws.cell(row=row, column=1, value='Talaba soni (Masofaviy jami)').alignment = Alignment(horizontal='center',
                                                                                           vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    # shakl background rangini och havo rang qil
    och_havo = PatternFill(start_color="b3dfe8", end_color="b3dfe8", fill_type="solid")
    ws.cell(row=row, column=1).fill = och_havo
    row += 1

    for k in [1, 2, 3, 4, 5, 6, 7]:
        jami_guruh = 0
        erkak_jami = 0
        ayol_jami = 0
        for guruh in bakalavr_guruhs:
            if k in kurs_set:
                budget = Budjet.objects.filter(guruhi=guruh, guruhi__kurs=k).aggregate(Sum('jami'), Sum('xotin_qiz'))
                shartnoma = Shartnoma.objects.filter(guruh=guruh, guruh__kurs=k).aggregate(Sum('jami'),
                                                                                           Sum('xotin_qiz'))
                if budget['jami__sum'] is not None:
                    jami_guruh += budget['jami__sum']
                    erkak_jami += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol_jami += budget['xotin_qiz__sum']
                    jami += budget['jami__sum']
                    erkak += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol += budget['xotin_qiz__sum']
                if shartnoma['jami__sum'] is not None:
                    jami_guruh += shartnoma['jami__sum']
                    erkak_jami += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol_jami += shartnoma['xotin_qiz__sum']
                    jami += shartnoma['jami__sum']
                    erkak += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol += shartnoma['xotin_qiz__sum']
        if jami_guruh != 0:
            ws.cell(row=row, column=1, value=k).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2, value=jami_guruh).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value=erkak_jami).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=ayol_jami).alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    ws.cell(row=row, column=1, value='Masofaviy jami').alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=jami).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=erkak).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=ayol).alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    kurs_set = Guruh.objects.filter(org=org, bosqich='Magistr').exclude(yonalish__mutahasislik_2=True).values_list(
        'kurs', flat=True).distinct()
    bakalavr_guruhs = Guruh.objects.filter(org=org, bosqich='Magistr').exclude(yonalish__mutahasislik_2=True).distinct()
    jami = 0
    erkak = 0
    ayol = 0
    ws.cell(row=row, column=1, value='Talaba soni (Magistr jami)').alignment = Alignment(horizontal='center',
                                                                                         vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    # shakl background rangini och havo rang qil
    och_havo = PatternFill(start_color="b3dfe8", end_color="b3dfe8", fill_type="solid")
    ws.cell(row=row, column=1).fill = och_havo
    row += 1

    for k in [1, 2, 3, 4, 5, 6, 7]:
        jami_guruh = 0
        erkak_jami = 0
        ayol_jami = 0
        for guruh in bakalavr_guruhs:
            if k in kurs_set:
                budget = Budjet.objects.filter(guruhi=guruh, guruhi__kurs=k).aggregate(Sum('jami'), Sum('xotin_qiz'))
                shartnoma = Shartnoma.objects.filter(guruh=guruh, guruh__kurs=k).aggregate(Sum('jami'),
                                                                                           Sum('xotin_qiz'))
                if budget['jami__sum'] is not None:
                    jami_guruh += budget['jami__sum']
                    erkak_jami += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol_jami += budget['xotin_qiz__sum']
                    jami += budget['jami__sum']
                    erkak += budget['jami__sum'] - budget['xotin_qiz__sum']
                    ayol += budget['xotin_qiz__sum']
                if shartnoma['jami__sum'] is not None:
                    jami_guruh += shartnoma['jami__sum']
                    erkak_jami += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol_jami += shartnoma['xotin_qiz__sum']
                    jami += shartnoma['jami__sum']
                    erkak += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
                    ayol += shartnoma['xotin_qiz__sum']
        if jami_guruh != 0:
            ws.cell(row=row, column=1, value=k).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2, value=jami_guruh).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value=erkak_jami).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=ayol_jami).alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    ws.cell(row=row, column=1, value='Magistr jami').alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=jami).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=erkak).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=ayol).alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    guruhs = Guruh.objects.filter(org=org)
    jami = 0
    erkak = 0
    ayol = 0

    row += 1
    for guruh in guruhs:
        budget = Budjet.objects.filter(guruhi=guruh).aggregate(Sum('jami'), Sum('xotin_qiz'))
        shartnoma = Shartnoma.objects.filter(guruh=guruh).aggregate(Sum('jami'), Sum('xotin_qiz'))
        if budget['jami__sum'] is not None:
            jami += budget['jami__sum']
            erkak += budget['jami__sum'] - budget['xotin_qiz__sum']
            ayol += budget['xotin_qiz__sum']
        if shartnoma['jami__sum'] is not None:
            jami += shartnoma['jami__sum']
            erkak += shartnoma['jami__sum'] - shartnoma['xotin_qiz__sum']
            ayol += shartnoma['xotin_qiz__sum']
    ws.cell(row=row, column=1, value='Jami').alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=jami).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=erkak).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=ayol).alignment = Alignment(horizontal='center', vertical='center')

    row += 2

    wb.save(output4)

    output4.seek(0)
    return output4

