import os
import re
from io import BytesIO

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()
from user.models import Faculty, Budjet, Shartnoma, Organization, Yonalish, Guruh

from datetime import datetime

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from openpyxl.utils import get_column_letter




organization_name = "Toshkent Davlat Texnnika Universiteti"




def exporttoexcell(org):
    def set_cell_properties(cell, value, alignment, font, border):
        cell.value = value
        cell.alignment = alignment
        cell.font = font
        cell.border = border
    FONT_NAME = 'Times New Roman'
    FONT_SIZE = 12
    FONT_COLOR = 'FF000000'
    BORDER_STYLE = 'thin'
    BORDER_COLOR = 'FF000000'
    red_color = 'FF0000FF'
    # color blue
    blue_color = 'FFFF0000'
    wb = Workbook()
    ws = wb.active
    output2 = BytesIO()
    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")
    organization_name = org.full_name
    ws.merge_cells('A1:P2')
    set_cell_properties(ws.cell(row=1, column=1),
                        f"{organization_name} talabalari kontingentining {formatted_time} holati haqida umumiy ma'lumot",
                        Alignment(horizontal='center', vertical='center'),
                        Font(name=FONT_NAME, size=12, bold=True, italic=False, color=FONT_COLOR),
                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

    cell_properties = [
        {'row': 3, 'column': 3, 'value': 'Guruhlar', 'width': True, 'merge': 'C3:C4'},
        {'row': 3, 'column': 4, 'value': 'Yo\'nalish', 'width': True, 'merge': 'D3:D4'},
        {'row': 3, 'column': 5, 'value': 'Jami', 'width': True, 'merge': 'E3:E4'},
        {'row': 3, 'column': 6, 'value': 'Shundan qizlar', 'width': True, 'merge': 'F3:F4'},
        {'row': 3, 'column': 7, 'value': 'Harbiy', 'width': True, 'merge': 'G3:G4'},
        {'row': 3, 'column': 8, 'value': 'Yangi qo\'shilgan', 'width': True, 'merge': 'H3:H4'},
        {'row': 3, 'column': 9, 'value': 'Chetlashtirilgan', 'width': True, 'merge': 'I3:I4'},
        {'row': 3, 'column': 10, 'value': 'Akademik tatil', 'width': True, 'merge': 'J3:J4'},
        {'row': 3, 'column': 11, 'value': 'Byudjet', 'width': True, 'merge': 'K3:M3'},
        {'row': 4, 'column': 11, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 12, 'value': 'Harbiy', 'width': True},
        {'row': 4, 'column': 13, 'value': 'Xotn qizlar', 'width': True},
        {'row': 3, 'column': 14, 'value': 'Shartmona', 'width': True, 'merge': 'N3:P3'},
        {'row': 4, 'column': 14, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 15, 'value': 'Harbiy', 'width': True},
        {'row': 4, 'column': 16, 'value': 'Xotn qizlar', 'width': True},

        # Add more dictionaries for other cells
    ]

    for properties in cell_properties:
        cell = ws.cell(row=properties['row'], column=properties['column'])
        set_cell_properties(cell,
                            properties['value'],
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
        if 'width' in properties and properties['width']:
            ws.column_dimensions[get_column_letter(properties['column'])].auto_size = True
        if 'merge' in properties:
            ws.merge_cells(properties['merge'])

    # d ustunni hammasiniblue qilish kerak FAQAT TEXTNI

    coluns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    for col in coluns:
        ws.column_dimensions[col].width = 17
    row = 5
    kurs = Guruh.objects.filter(org=org).values('kurs').distinct()
    for k in kurs:
        ws.merge_cells(f'C{row}:P{row}')
        set_cell_properties(ws.cell(row=row, column=3),
                            # k['kurs'],
                            f'{k["kurs"]}-kurs',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

        row += 1
        faculty = Faculty.objects.filter(org=org, org__guruh__kurs=k['kurs'], org__guruh__isnull=False).distinct()
        for f in faculty:
            yonalish = Yonalish.objects.filter(faculty=f, yonalishguruh__kurs=k['kurs']).distinct()
            for y in yonalish:
                guruh = Guruh.objects.filter(yonalish=y, org=org, kurs=k['kurs']).distinct()
                for g in guruh:
                    # faqat jamini ol hisoblama
                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'),
                                                                            Sum('xotin_qiz'))

                    jami = (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
                    shundan_qizlar = (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
                    harbiy = (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
                    yangi_qoshilgan = g.new_students

                    chetlashtirilgan = g.chetlashtirilgan_students
                    akademik_tatil = g.akademik
                    yname2 = ''.join(re.findall(r'\b\w', y.name)).upper()
                    set_cell_properties(ws.cell(row=row, column=3), g.name,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    # set_cell_properties(ws.cell(row=row, column=4), y.name,
                    #                     Alignment(horizontal='center', vertical='center'),
                    #                     Font(name=FONT_NAME, size=FONT_SIZE,bold=False, italic=False, color=FONT_COLOR),
                    #                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                    #                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                    #                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                    #                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    if y.language == 'O\'zbek':
                        set_cell_properties(ws.cell(row=row, column=4), f'{y.name} o`z',
                                            Alignment(horizontal='center', vertical='center'),
                                            Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                                 color=FONT_COLOR),
                                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    else:
                        set_cell_properties(ws.cell(row=row, column=4), f'{yname2} ru',
                                            Alignment(horizontal='center', vertical='center'),
                                            Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                                 color=FONT_COLOR),
                                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                    set_cell_properties(ws.cell(row=row, column=5), jami,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    # red for text

                    set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=blue_color),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=11), budjet['jami__sum'],
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=12), budjet['harbiy__sum'],
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=13), budjet['xotin_qiz__sum'],
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=14), shartnoma['jami__sum'],
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=15), shartnoma['harbiy__sum'],
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                    set_cell_properties(ws.cell(row=row, column=16), shartnoma['xotin_qiz__sum'],
                                        Alignment(horizontal='center', vertical='center'),
                                        Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False,
                                             color=FONT_COLOR),
                                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                    row += 1

                #             manashu yonalishdagi hamma ustunlar bo'yicha jami sini chiqarihs kerak
                #             keyin boshqa yonalishga o'tadi
                #         keyin boshqa fakultetga o'tadi
                #     keyin boshqa kursga o'tadi
                y_name = y.name
                y_name2 = ''.join(re.findall(r'\b\w', y_name)).upper()

                # manashu yonalishga bog'langan guruhlarni  jami hisobla
                y_full_jami = y.yonalishguruh.filter(org=org, kurs=k['kurs']).aggregate(Sum('guruhbudjet__jami'),
                                                                                        Sum('guruhbudjet__harbiy'),
                                                                                        Sum('guruhbudjet__xotin_qiz'),
                                                                                        Sum('guruhshartnoma__jami'),
                                                                                        Sum('guruhshartnoma__harbiy'),
                                                                                        Sum('guruhshartnoma__xotin_qiz'))

                new_student_sum = y.yonalishguruh.filter(org=org, kurs=k['kurs']).aggregate(Sum('new_students'))
                chetlashtirilgan_student_sum = y.yonalishguruh.filter(org=org, kurs=k['kurs']).aggregate(
                    Sum('chetlashtirilgan_students'))
                akademik_sum = y.yonalishguruh.filter(org=org, kurs=k['kurs']).aggregate(Sum('akademik'))
                set_cell_properties(ws.cell(row=row, column=3), f'',
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                set_cell_properties(ws.cell(row=row, column=4), f'{y_name2}-jami',
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=5),
                                    # (y_full_jami['budjet__jami__sum'] or 0) + (
                                    #         y_full_jami['shartnoma__jami__sum'] or 0),
                                    (y_full_jami['guruhbudjet__jami__sum'] or 0) + (
                                            y_full_jami['guruhshartnoma__jami__sum'] or 0),
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=6),
                                    (y_full_jami['guruhbudjet__xotin_qiz__sum'] or 0) + (
                                            y_full_jami['guruhshartnoma__xotin_qiz__sum'] or 0),
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=7),
                                    (y_full_jami['guruhbudjet__harbiy__sum'] or 0) + (
                                            y_full_jami['guruhshartnoma__harbiy__sum'] or 0),
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=8), new_student_sum['new_students__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=9),
                                    chetlashtirilgan_student_sum['chetlashtirilgan_students__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=10), akademik_sum['akademik__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=11), y_full_jami['guruhbudjet__jami__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=12), y_full_jami['guruhbudjet__harbiy__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=13), y_full_jami['guruhbudjet__xotin_qiz__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=14), y_full_jami['guruhshartnoma__jami__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=15), y_full_jami['guruhshartnoma__harbiy__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                set_cell_properties(ws.cell(row=row, column=16), y_full_jami['guruhshartnoma__xotin_qiz__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

                row += 1
            #            fakultetga bog'langan yonalishlarni jami hisobla
            f_name = f.name

            f_turi = Yonalish.objects.filter(faculty=f, yonalishguruh__kurs=k['kurs']).values('turi')

            f_full_jami = Guruh.objects.filter(
                org=org,
                yonalish__faculty=f,
                kurs=k['kurs']).aggregate(Sum('guruhbudjet__jami'),
                                          Sum('guruhbudjet__harbiy'),
                                          Sum('guruhbudjet__xotin_qiz'),
                                          Sum('guruhshartnoma__jami'),
                                          Sum('guruhshartnoma__harbiy'),
                                          Sum('guruhshartnoma__xotin_qiz'))

            new_student_sum = Guruh.objects.filter(
                org=org,
                yonalish__faculty=f,
                kurs=k['kurs']
            ).aggregate(Sum('new_students'))

            chetlashtirilgan_student_sum = Guruh.objects.filter(
                org=org,
                yonalish__faculty=f,
                kurs=k['kurs']
            ).aggregate(Sum('chetlashtirilgan_students'))

            akademik_sum = Guruh.objects.filter(
                org=org,
                yonalish__faculty=f,
                kurs=k['kurs']
            ).aggregate(Sum('akademik'))

            jami = (f_full_jami['guruhbudjet__jami__sum'] or 0) + (f_full_jami['guruhshartnoma__jami__sum'] or 0)
            shundan_qizlar = (f_full_jami['guruhbudjet__xotin_qiz__sum'] or 0) + (
                    f_full_jami['guruhshartnoma__xotin_qiz__sum'] or 0)
            harbiy = (f_full_jami['guruhbudjet__harbiy__sum'] or 0) + (f_full_jami['guruhshartnoma__harbiy__sum'] or 0)
            yangi_qoshilgan = new_student_sum['new_students__sum']
            chetlashtirilgan = chetlashtirilgan_student_sum['chetlashtirilgan_students__sum']
            akademik_tatil = akademik_sum['akademik__sum']
            # fakultetning nomida qatnashgan so;zlarni faqat bosh harifini olish kerak

            # f_name = re.findall(r'\b\w', f_name)

            f_bosqich = Guruh.objects.filter(
                org=org,
                yonalish__faculty=f,
                kurs=k['kurs']
            ).values('bosqich')

            f_name2 = ''
            # xullas magistratura sirtqi va kunduzgi masofaviy bo'limlarini topish kerak
            # <QuerySet [{'turi': 'Masofaviy'}]>

            f_bosqich = ''.join([i['bosqich'] for i in f_bosqich]).split(' ')

            f_turi = ''.join([i['turi'] for i in f_turi])
            if f_turi == 'Masofaviy':
                f_name = 'Masofaviy' + '' + f_name
            if f_turi == 'Sirtqi':
                f_name = f_name + ' Sirtqi'

            if f_bosqich == 'Magistr':
                f_name = f_name + ' Magistr'

            if jami != 0:
                # set_cell_properties(ws.cell(row=row, column=3), f'',
                #                     Alignment(horizontal='center', vertical='center'),
                #                     Font(name=FONT_NAME, size=FONT_SIZE,bold=False, italic=False, color=FONT_COLOR),
                #                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                #                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                #                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                #                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                # set_cell_properties(ws.cell(row=row, column=4), f'{f_name} JAMI',
                #                     Alignment(horizontal='center', vertical='center'),
                #                     Font(name=FONT_NAME, size=FONT_SIZE,bold=False, italic=False, color=FONT_COLOR),
                #                     Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                #                            right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                #                            top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                #                            bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                # use merge ws.merge_cells(f'C{row}:D{row}')
                ws.merge_cells(f'C{row}:D{row}')
                set_cell_properties(ws.cell(row=row, column=3), f_name,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR))
                # color pink bor box
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                              fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=5), jami,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                              fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=blue_color),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                              fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                              fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                              fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                              fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=11), f_full_jami['guruhbudjet__jami__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=12), f_full_jami['guruhbudjet__harbiy__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=13), f_full_jami['guruhbudjet__xotin_qiz__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=14), f_full_jami['guruhshartnoma__jami__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=15), f_full_jami['guruhshartnoma__harbiy__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")
                set_cell_properties(ws.cell(row=row, column=16), f_full_jami['guruhshartnoma__xotin_qiz__sum'],
                                    Alignment(horizontal='center', vertical='center'),
                                    Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                    Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                           bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
                ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                               fill_type="solid")

                row += 1
        gurhlar = Guruh.objects.filter(org=org, kurs=k['kurs'])
        guruhlar2 = gurhlar.filter(bosqich='Magistr')
        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        for g in guruhlar2:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami = budjet['jami__sum'] or 0
            shartnoma_jami = shartnoma['jami__sum'] or 0
            budjet_harbiy = budjet['harbiy__sum'] or 0
            shartnoma_harbiy = shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz = budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz = shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), 'MAGISTRATURA',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=blue_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1
        sirtqi_jamis = Guruh.objects.filter(
            org=org,
            yonalish__turi='Sirtqi',
            kurs=k['kurs'])
        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        for g in sirtqi_jamis:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), 'SIRTQI',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=blue_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1
        doktorantura = Guruh.objects.filter(
            org=org,
            yonalish__turi='Doktorantura',
            kurs=k['kurs'])
        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        #     och ko'k rangdan foydalan
        for g in doktorantura:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), 'DOKTORANTURA',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=blue_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row
                    =row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                       fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1
        kunduzgi = Guruh.objects.filter(
            org=org,
            yonalish__turi='Kunduzgi',
            kurs=k['kurs'])
        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        for g in kunduzgi:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), 'KUNDUZGI',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=blue_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1
        masofaviy = Guruh.objects.filter(
            org=org,
            yonalish__turi='Masofaviy',
            kurs=k['kurs'])

        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        for g in masofaviy:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), 'MASOFAVIY',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=blue_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=False, color=FONT_COLOR),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1

        mutahasislik_2_jami = Guruh.objects.filter(
            org=org,
            kurs=k['kurs'],
            yonalish__mutahasislik_2=True
        )
        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        for g in mutahasislik_2_jami:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), '2-MUTAHASISLIK JAMI',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1

        jami_kurs = Guruh.objects.filter(
            org=org,
            kurs=k['kurs'])
        jami = 0
        shundan_qizlar = 0
        harbiy = 0
        yangi_qoshilgan = 0
        chetlashtirilgan = 0
        akademik_tatil = 0
        budjet_jami = 0
        shartnoma_jami = 0
        budjet_harbiy = 0
        shartnoma_harbiy = 0
        budjet_xotin_qiz = 0
        shartnoma_xotin_qiz = 0

        for g in jami_kurs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

        if jami != 0:
            ws.merge_cells(f'C{row}:D{row}')
            set_cell_properties(ws.cell(row=row, column=3), f'{k["kurs"]}-KURS Jami',
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

            ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=5), jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

            ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=7), harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=7).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

            ws.cell(row=row, column=8).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=9).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                          fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=11).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=12).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=13).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

            ws.cell(row=row, column=14).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=15).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")
            set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                                Alignment(horizontal='center', vertical='center'),
                                Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                                Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                       bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
            ws.cell(row=row, column=16).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB",
                                                           fill_type="solid")

            row += 1

    ws.merge_cells(f'C{row}:P{row}')
    #     to'q qizil
    ws.cell(row=row, column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                  fill_type="solid")
    row += 1

    fakultetlar_kunduzgi = Faculty.objects.filter(org=org, facultyyonalish__turi='Kunduzgi').distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org).distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'KUNDUZGI JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org, facultyyonalish__turi='Sirtqi').distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org).distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'SIRQI JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1

    fakultetlar_kunduzgi = Faculty.objects.filter(org=org, facultyyonalish__turi='Masofaviy').distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org).distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'MASOFAVIY JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org, facultyyonalish__mutahasislik_2=True).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org).distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), '2-MUTAHASISLIK JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Bakalavr').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'BAKALVR JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Magistr').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'MAGISTR JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Doktorantura').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'DOKTORANTURA JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Bakalavr',
                                      yonalish__turi='Kunduzgi').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'BAKALAVR KUNDUZGI JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Bakalavr',
                                      yonalish__turi='Sirtqi').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'BAKALAVR SIRTQI  JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Bakalavr',
                                      yonalish__turi='Masofaviy').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'BAKALAVR MASOFAVIY JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Magistr',
                                      yonalish__turi='Kunduzgi').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'MAGISTIR KUNDUZGI  JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Magistr',
                                      yonalish__turi='Sirtqi').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'MAGISTIR SIRTQI  JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Magistr',
                                      yonalish__turi='Masofaviy').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'MAGISTR MASOFAVIY  JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Doktorantura',
                                      yonalish__turi='Kunduzgi').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'DOKTORANTURA KUNDUZGI JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Doktorantura',
                                      yonalish__turi='Sirtqi').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'DOKTORANTURA SIRTQI JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    for f in fakultetlar_kunduzgi:
        guruhs = Guruh.objects.filter(yonalish__faculty=f, org=org, bosqich='Doktorantura',
                                      yonalish__turi='Masofaviy').distinct()
        for g in guruhs:
            budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
            jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
            shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
            harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
            yangi_qoshilgan += g.new_students
            chetlashtirilgan += g.chetlashtirilgan_students
            akademik_tatil += g.akademik
            budjet_jami += budjet['jami__sum'] or 0
            shartnoma_jami += shartnoma['jami__sum'] or 0
            budjet_harbiy += budjet['harbiy__sum'] or 0
            shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
            budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
            shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), 'DOKTORANTURA MASOFAVIY JAMI',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1
    fakultetlar_kunduzgi = Faculty.objects.filter(org=org).distinct()

    jami = 0
    shundan_qizlar = 0
    harbiy = 0
    yangi_qoshilgan = 0
    chetlashtirilgan = 0
    akademik_tatil = 0
    budjet_jami = 0
    shartnoma_jami = 0
    budjet_harbiy = 0
    shartnoma_harbiy = 0
    budjet_xotin_qiz = 0
    shartnoma_xotin_qiz = 0

    guruhs = Guruh.objects.filter(org=org)
    for g in guruhs:
        budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
        shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'), Sum('harbiy'), Sum('xotin_qiz'))
        jami += (budjet['jami__sum'] or 0) + (shartnoma['jami__sum'] or 0)
        shundan_qizlar += (budjet['xotin_qiz__sum'] or 0) + (shartnoma['xotin_qiz__sum'] or 0)
        harbiy += (budjet['harbiy__sum'] or 0) + (shartnoma['harbiy__sum'] or 0)
        yangi_qoshilgan += g.new_students
        chetlashtirilgan += g.chetlashtirilgan_students
        akademik_tatil += g.akademik
        budjet_jami += budjet['jami__sum'] or 0
        shartnoma_jami += shartnoma['jami__sum'] or 0
        budjet_harbiy += budjet['harbiy__sum'] or 0
        shartnoma_harbiy += shartnoma['harbiy__sum'] or 0
        budjet_xotin_qiz += budjet['xotin_qiz__sum'] or 0
        shartnoma_xotin_qiz += shartnoma['xotin_qiz__sum'] or 0

    if jami != 0:
        #     och havo rang
        ws.merge_cells(f'C{row}:D{row}')
        set_cell_properties(ws.cell(row=row, column=3), ' JAMI TALABALAR SONI ',
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=4).border = Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=5), jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=6), shundan_qizlar,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=7), harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=8), yangi_qoshilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=8).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=9), chetlashtirilgan,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=9).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                      fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=10), akademik_tatil,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=10).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=11), budjet_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=11).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=12), budjet_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=12).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=13), budjet_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=blue_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=13).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")

        set_cell_properties(ws.cell(row=row, column=14), shartnoma_jami,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))

        ws.cell(row=row, column=14).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=15), shartnoma_harbiy,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=15).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        set_cell_properties(ws.cell(row=row, column=16), shartnoma_xotin_qiz,
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=red_color),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=FONT_COLOR)))
        ws.cell(row=row, column=16).fill = PatternFill(start_color="87CEFA", end_color="87CEFA",
                                                       fill_type="solid")
        row += 1

    wb.save(output2)
    output2.seek(0)
    return output2
    # wb.save('talabalar.xlsx')
