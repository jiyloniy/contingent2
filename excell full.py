import os
from io import BytesIO
import django
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from ecxel2 import exporttoexcel
from exportexcel import exporttoexcell
from user.models import Organization

org = Organization.objects.filter(name='kiuf').first()

def generate_excel_files(org):
    output2 = exporttoexcel(org)
    output1 = exporttoexcell(org)
    return [output1, output2]

def copy_cell_style(source_cell, dest_cell):
    dest_cell.font = copy_font(source_cell.font)
    dest_cell.border = copy_border(source_cell.border)
    dest_cell.fill = copy_fill(source_cell.fill)
    dest_cell.alignment = copy_alignment(source_cell.alignment)

def copy_font(font):
    return Font(name=font.name, size=font.size, bold=font.bold, italic=font.italic, color=font.color)

def copy_border(border):
    if border is None:
        return None
    return Border(left=copy_side(border.left), right=copy_side(border.right),
                  top=copy_side(border.top), bottom=copy_side(border.bottom))

def copy_side(side):
    if side is None:
        return None
    return Side(border_style=side.border_style, color=side.color)

def copy_fill(fill):
    return PatternFill(fill_type=fill.fill_type, start_color=fill.start_color, end_color=fill.end_color)

def copy_alignment(alignment):
    return Alignment(horizontal=alignment.horizontal, vertical=alignment.vertical)

def merge_excel_files(excel_files):
    merged_output = BytesIO()
    merged_workbook = openpyxl.Workbook()

    # Remove the default sheet created
    default_sheet = merged_workbook.active
    merged_workbook.remove(default_sheet)

    for index, file_bytes in enumerate(excel_files, start=1):
        workbook = openpyxl.load_workbook(file_bytes)
        worksheet = workbook.active

        sheet_name = f"Sheet{index}"
        merged_workbook.create_sheet(title=sheet_name)
        merged_sheet = merged_workbook[sheet_name]

        for row in worksheet.iter_rows():
            for cell in row:
                merged_cell = merged_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                copy_cell_style(cell, merged_cell)

        # Merge cells if specified
        merge_cells = worksheet.merged_cells.ranges
        for merge_range in merge_cells:
            merged_sheet.merge_cells(merge_range.coord)

    merged_workbook.save('talabalar.xlsx')


# Generate Excel files
excel_files = generate_excel_files(org)

# Merge Excel files
merge_excel_files(excel_files)
