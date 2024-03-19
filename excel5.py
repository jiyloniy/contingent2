import os

from io import BytesIO

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()

from datetime import datetime

from django.db.models import Sum

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

django.setup()
from user.models import Faculty, Budjet, Shartnoma, Organization, Yonalish, Guruh



def exporttoexcel5(org):
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl import Workbook

    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")

    def set_cell_properties(cell, value, alignment, font, border):
        cell.value = value
        cell.alignment = alignment
        cell.font = font
        cell.border = border

    FONT_NAME = 'Times New Roman'
    FONT_SIZE = 8
    FONT_COLOR = 'FF000000'
    BORDER_STYLE = 'thin'
    BORDER_COLOR = 'FF000000'
    red_color = 'FF0000FF'
    # color blue
    blue_color = 'FFFF0000'

    # add auto size width of column

    # Hozirgi vaqtni ol
    # add auto size width of column

    # Hozirgi vaqtni olib, formatini belgilash

    now = datetime.now()
    formatted_time = now.strftime("%Y-%m-%d")
    wb = Workbook()
    ws = wb.active
    output4 = BytesIO()
    organization_name = org.full_name

    ws.merge_cells('A1:AD2')
    set_cell_properties(ws.cell(row=1, column=1),
                        f"{organization_name} talabalari kontingentining {formatted_time} holati haqida umumiy ma'lumot (Davlat granti/To'lov-kontrakt)",
                        Alignment(horizontal='center', vertical='center'),
                        Font(name=FONT_NAME, size=18, bold=True, italic=False, color=FONT_COLOR),
                        Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                               bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))

    cell_properties = [
        {'row': 3, 'column': 1, 'value': '№', 'width': True, 'merge': 'A3:A4'},
        {'row': 3, 'column': 2, 'value': 'Ta\'lim yo\'nalishi kodi va nomi', 'width': True, 'merge': 'B3:B4'},
        {'row': 3, 'column': 3, 'value': 'Ta\'lim turi', 'width': True, 'merge': 'C3:C4'},
        {'row': 3, 'column': 4, 'value': 'Jami', 'width': True, 'merge': 'D3:D4'},
        {'row': 3, 'column': 5, 'value': 'Jami', 'width': True, 'merge': 'E3:F3'},
        {'row': 4, 'column': 5, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 6, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 7, 'value': '1-kurs', 'width': True, 'merge': 'G3:I3'},
        {'row': 4, 'column': 7, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 8, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 9, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 10, 'value': '2-kurs', 'width': True, 'merge': 'J3:L3'},
        {'row': 4, 'column': 10, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 11, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 12, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 13, 'value': '3-kurs', 'width': True, 'merge': 'M3:O3'},
        {'row': 4, 'column': 13, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 14, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 15, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 16, 'value': '4-kurs', 'width': True, 'merge': 'P3:R3'},
        {'row': 4, 'column': 16, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 17, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 18, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 19, 'value': '4-kurs', 'width': True, 'merge': 'S3:U3'},
        {'row': 4, 'column': 19, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 20, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 21, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 22, 'value': '5-kurs', 'width': True, 'merge': 'V3:X3'},
        {'row': 4, 'column': 22, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 23, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 24, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 25, 'value': '5-kurs', 'width': True, 'merge': 'Y3:AA3'},
        {'row': 4, 'column': 25, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 26, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 27, 'value': 'To\'lov-kontrakt', 'width': True},
        {'row': 3, 'column': 28, 'value': '6-kurs', 'width': True, 'merge': 'AB3:AD3'},
        {'row': 4, 'column': 28, 'value': 'Jami', 'width': True},
        {'row': 4, 'column': 29, 'value': 'Davlat granti', 'width': True},
        {'row': 4, 'column': 30, 'value': 'To\'lov-kontrakt', 'width': True},
    ]

    for properties in cell_properties:
        cell = ws.cell(row=properties['row'], column=properties['column'])
        set_cell_properties(cell,
                            properties['value'],
                            Alignment(horizontal='center', vertical='center'),
                            Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=False, color=FONT_COLOR),
                            Border(left=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   right=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   top=Side(border_style=BORDER_STYLE, color=BORDER_COLOR),
                                   bottom=Side(border_style=BORDER_STYLE, color=BORDER_COLOR)))
        # if 'width' in properties and properties['width']:
        #     ws.column_dimensions[get_column_letter(properties['column'])].auto_size = True
        if 'merge' in properties:
            ws.merge_cells(properties['merge'])

        faculty = Faculty.objects.filter(org=org, org__guruh__isnull=False).distinct()

        '''
        xullas  menga shunday qilish kerak oldin facultetlarni oladi va xar bir fakultet bo'yicha yo'nlasihlarni filtert qiladi bu filter Kunduzgi Sirtqi Masofaviy Magistr Doktorant mutahasislik_2=True bo'yicha bo'ladi bu yo'nalsihlar tugagandan keyin har bir filter bo'yicha jamini hsipb;ash kerak undan song har bir yo'nalishga bo'glangan guruhlarlarni oladi guruhlar kurslar kesimi bo'yicha filter qilinadi masalan 1- kurs bo'lgan hamma guruh 2- kurs bo'lgan hamma guruhlarni oladi hisoblaydi davlat grandi Budget modelidagi jami ga teng Shartnoma esa Shartnoma jami ga teng jami 2 tasini yig'indisaga teng bo'ladi har bir fakultet tugagandan keyin 
        '''
        row = 5

        jami2_full = 0
        jami2_b = 0
        jami2_c = 0
        jami2_1_kurs = 0
        jami2_1_kurs_b = 0
        jami2_1_kurs_c = 0
        jami2_2_kurs = 0
        jami2_2_kurs_b = 0
        jami2_2_kurs_c = 0
        jami2_3_kurs = 0
        jami2_3_kurs_b = 0
        jami2_3_kurs_c = 0
        jami2_4_kurs = 0
        jami2_4_kurs_b = 0
        jami2_4_kurs_c = 0
        jami2_5_kurs = 0
        jami2_5_kurs_b = 0
        jami2_5_kurs_c = 0
        jami2_6_kurs = 0
        jami2_6_kurs_b = 0
        jami2_6_kurs_c = 0
        jami2_7_kurs = 0
        jami2_7_kurs_b = 0
        jami2_7_kurs_c = 0
        for f in faculty:
            jami_full = 0
            jami_b = 0
            jami_c = 0
            jami_1_kurs = 0
            jami_1_kurs_b = 0
            jami_1_kurs_c = 0
            jami_2_kurs = 0
            jami_2_kurs_b = 0
            jami_2_kurs_c = 0
            jami_3_kurs = 0
            jami_3_kurs_b = 0
            jami_3_kurs_c = 0
            jami_4_kurs = 0
            jami_4_kurs_b = 0
            jami_4_kurs_c = 0
            jami_5_kurs = 0
            jami_5_kurs_b = 0
            jami_5_kurs_c = 0
            jami_6_kurs = 0
            jami_6_kurs_b = 0
            jami_6_kurs_c = 0
            jami_7_kurs = 0
            jami_7_kurs_b = 0
            jami_7_kurs_c = 0

            yonalish_kunduzgi = Yonalish.objects.filter(faculty=f, turi='Kunduzgi').exclude(yonalishguruh__bosqich='Magistr').exclude(
                yonalishguruh__bosqich='Doktorant')
            yonalish_Sirtqi = Yonalish.objects.filter(faculty=f, turi='Sirtqi')
            yonalish_Masofaviy = Yonalish.objects.filter(faculty=f, turi='Masofaviy')
            yonalish_Magistratura = Yonalish.objects.filter(faculty=f, yonalishguruh__bosqich='Magistr')
            yonalish_Doktorantura = Yonalish.objects.filter(faculty=f, yonalishguruh__bosqich='Doktorant')
            yonalish_ikkinchi_talim = Yonalish.objects.filter(faculty=f, mutahasislik_2=True)

            for y in yonalish_kunduzgi:

                jami = 0
                jami_b = 0
                jami_c = 0
                yonalish_jami_1_kurs = 0
                yonalish_jami_1_kurs_b = 0
                yonalish_jami_1_kurs_c = 0
                yonalish_jami_2_kurs = 0
                yonalish_jami_2_kurs_b = 0
                yonalish_jami_2_kurs_c = 0
                yonalish_jami_3_kurs = 0
                yonalish_jami_3_kurs_b = 0
                yonalish_jami_3_kurs_c = 0
                yonalish_jami_4_kurs = 0
                yonalish_jami_4_kurs_b = 0
                yonalish_jami_4_kurs_c = 0
                yonalish_jami_5_kurs = 0
                yonalish_jami_5_kurs_b = 0
                yonalish_jami_5_kurs_c = 0
                yonalish_jami_6_kurs = 0
                yonalish_jami_6_kurs_b = 0
                yonalish_jami_6_kurs_c = 0
                yonalish_jami_7_kurs = 0
                yonalish_jami_7_kurs_b = 0
                yonalish_jami_7_kurs_c = 0
                guruhs = Guruh.objects.filter(org=org, yonalish=y, yonalish__faculty=f)
                for g in guruhs:

                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'))
                    if budjet['jami__sum']:
                        jami_full += budjet['jami__sum']
                        jami2_full += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']
                        jami += budjet['jami__sum']
                        jami_b += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']

                    if shartnoma['jami__sum']:
                        jami_full += shartnoma['jami__sum']
                        jami2_full += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                        jami += shartnoma['jami__sum']
                        jami_c += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                    if jami != 0:
                        if g.kurs == 1:
                            if budjet['jami__sum']:
                                jami_1_kurs += budjet['jami__sum']
                                jami2_1_kurs += budjet['jami__sum']
                                jami2_1_kurs_b += budjet['jami__sum']
                                jami_1_kurs_b += budjet['jami__sum']
                                yonalish_jami_1_kurs += budjet['jami__sum']
                                yonalish_jami_1_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs_c += shartnoma['jami__sum']
                                jami_1_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_1_kurs += shartnoma['jami__sum']
                                yonalish_jami_1_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 2:
                            if budjet['jami__sum']:
                                jami_2_kurs += budjet['jami__sum']
                                jami2_2_kurs += budjet['jami__sum']
                                jami2_2_kurs_b += budjet['jami__sum']
                                jami_2_kurs_b += budjet['jami__sum']
                                yonalish_jami_2_kurs += budjet['jami__sum']
                                yonalish_jami_2_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs_c += shartnoma['jami__sum']
                                jami_2_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_2_kurs += shartnoma['jami__sum']
                                yonalish_jami_2_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 3:
                            if budjet['jami__sum']:
                                jami_3_kurs += budjet['jami__sum']
                                jami2_3_kurs += budjet['jami__sum']
                                jami2_3_kurs_b += budjet['jami__sum']
                                jami_3_kurs_b += budjet['jami__sum']
                                yonalish_jami_3_kurs += budjet['jami__sum']
                                yonalish_jami_3_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_3_kurs += shartnoma['jami__sum']
                                jami_3_kurs_c += shartnoma['jami__sum']
                                jami2_3_kurs += shartnoma['jami__sum']
                                jami2_3_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_3_kurs += shartnoma['jami__sum']
                                yonalish_jami_3_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 4:
                            if budjet['jami__sum']:
                                jami_4_kurs += budjet['jami__sum']
                                jami_4_kurs_b += budjet['jami__sum']
                                jami2_4_kurs += budjet['jami__sum']
                                jami2_4_kurs_b += budjet['jami__sum']
                                yonalish_jami_4_kurs += budjet['jami__sum']
                                yonalish_jami_4_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs_c += shartnoma['jami__sum']
                                jami_4_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_4_kurs += shartnoma['jami__sum']
                                yonalish_jami_4_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 5:
                            if budjet['jami__sum']:
                                jami_5_kurs += budjet['jami__sum']
                                jami2_5_kurs += budjet['jami__sum']
                                jami2_5_kurs_b += budjet['jami__sum']
                                jami_5_kurs_b += budjet['jami__sum']
                                yonalish_jami_5_kurs += budjet['jami__sum']
                                yonalish_jami_5_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_5_kurs += shartnoma['jami__sum']
                                jami_5_kurs_c += shartnoma['jami__sum']
                                jami2_5_kurs += shartnoma['jami__sum']
                                jami2_5_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_5_kurs += shartnoma['jami__sum']
                                yonalish_jami_5_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 6:
                            if budjet['jami__sum']:
                                jami_6_kurs += budjet['jami__sum']
                                jami_6_kurs_b += budjet['jami__sum']
                                jami2_6_kurs += budjet['jami__sum']
                                jami2_6_kurs_b += budjet['jami__sum']
                                yonalish_jami_6_kurs += budjet['jami__sum']
                                yonalish_jami_6_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs_c += shartnoma['jami__sum']
                                jami_6_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_6_kurs += shartnoma['jami__sum']
                                yonalish_jami_6_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 7:
                            if budjet['jami__sum']:
                                jami_7_kurs += budjet['jami__sum']
                                jami2_7_kurs += budjet['jami__sum']
                                jami_7_kurs_b += budjet['jami__sum']
                                yonalish_jami_7_kurs += budjet['jami__sum']
                                yonalish_jami_7_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_7_kurs += shartnoma['jami__sum']
                                jami2_7_kurs += shartnoma['jami__sum']
                                jami_7_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_7_kurs += shartnoma['jami__sum']
                                yonalish_jami_7_kurs_c += shartnoma['jami__sum']


                if jami != 0:
                    ws.cell(row=row, column=2, value=y.name)
                    ws.cell(row=row, column=3, value=y.turi)
                    ws.cell(row=row, column=4, value=jami)
                    ws.cell(row=row, column=5, value=jami_b)
                    ws.cell(row=row, column=6, value=jami_c)
                    ws.cell(row=row, column=7, value=yonalish_jami_1_kurs)
                    ws.cell(row=row, column=8, value=yonalish_jami_1_kurs_b)
                    ws.cell(row=row, column=9, value=yonalish_jami_1_kurs_c)
                    ws.cell(row=row, column=10, value=yonalish_jami_2_kurs)
                    ws.cell(row=row, column=11, value=yonalish_jami_2_kurs_b)
                    ws.cell(row=row, column=12, value=yonalish_jami_2_kurs_c)
                    ws.cell(row=row, column=13, value=yonalish_jami_3_kurs)
                    ws.cell(row=row, column=14, value=yonalish_jami_3_kurs_b)
                    ws.cell(row=row, column=15, value=yonalish_jami_3_kurs_c)
                    ws.cell(row=row, column=16, value=yonalish_jami_4_kurs)
                    ws.cell(row=row, column=17, value=yonalish_jami_4_kurs_b)
                    ws.cell(row=row, column=18, value=yonalish_jami_4_kurs_c)
                    ws.cell(row=row, column=19, value=yonalish_jami_5_kurs)
                    ws.cell(row=row, column=20, value=yonalish_jami_5_kurs_b)
                    ws.cell(row=row, column=21, value=yonalish_jami_5_kurs_c)
                    ws.cell(row=row, column=22, value=yonalish_jami_6_kurs)
                    ws.cell(row=row, column=23, value=yonalish_jami_6_kurs_b)
                    ws.cell(row=row, column=24, value=yonalish_jami_6_kurs_c)
                    ws.cell(row=row, column=25, value=yonalish_jami_7_kurs)
                    ws.cell(row=row, column=26, value=yonalish_jami_7_kurs_b)
                    ws.cell(row=row, column=27, value=yonalish_jami_7_kurs_c)
                    row += 1
            for y in yonalish_Magistratura:

                jami = 0
                jami_b = 0
                jami_c = 0
                yonalish_jami_1_kurs = 0
                yonalish_jami_1_kurs_b = 0
                yonalish_jami_1_kurs_c = 0
                yonalish_jami_2_kurs = 0
                yonalish_jami_2_kurs_b = 0
                yonalish_jami_2_kurs_c = 0
                yonalish_jami_3_kurs = 0
                yonalish_jami_3_kurs_b = 0
                yonalish_jami_3_kurs_c = 0
                yonalish_jami_4_kurs = 0
                yonalish_jami_4_kurs_b = 0
                yonalish_jami_4_kurs_c = 0
                yonalish_jami_5_kurs = 0
                yonalish_jami_5_kurs_b = 0
                yonalish_jami_5_kurs_c = 0
                yonalish_jami_6_kurs = 0
                yonalish_jami_6_kurs_b = 0
                yonalish_jami_6_kurs_c = 0
                yonalish_jami_7_kurs = 0
                yonalish_jami_7_kurs_b = 0
                yonalish_jami_7_kurs_c = 0
                guruhs = Guruh.objects.filter(org=org, yonalish=y, yonalish__faculty=f)
                for g in guruhs:

                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'))
                    if budjet['jami__sum']:
                        jami_full += budjet['jami__sum']
                        jami2_full += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']
                        jami += budjet['jami__sum']
                        jami_b += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']

                    if shartnoma['jami__sum']:
                        jami_full += shartnoma['jami__sum']
                        jami2_full += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                        jami += shartnoma['jami__sum']
                        jami_c += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                    if jami != 0:
                        if g.kurs == 1:
                            if budjet['jami__sum']:
                                jami_1_kurs += budjet['jami__sum']
                                jami2_1_kurs += budjet['jami__sum']
                                jami2_1_kurs_b += budjet['jami__sum']
                                jami_1_kurs_b += budjet['jami__sum']
                                yonalish_jami_1_kurs += budjet['jami__sum']
                                yonalish_jami_1_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs_c += shartnoma['jami__sum']
                                jami_1_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_1_kurs += shartnoma['jami__sum']
                                yonalish_jami_1_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 2:
                            if budjet['jami__sum']:
                                jami_2_kurs += budjet['jami__sum']
                                jami2_2_kurs += budjet['jami__sum']
                                jami2_2_kurs_b += budjet['jami__sum']
                                jami_2_kurs_b += budjet['jami__sum']
                                yonalish_jami_2_kurs += budjet['jami__sum']
                                yonalish_jami_2_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs_c += shartnoma['jami__sum']
                                jami_2_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_2_kurs += shartnoma['jami__sum']
                                yonalish_jami_2_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 3:
                            if budjet['jami__sum']:
                                jami_3_kurs += budjet['jami__sum']
                                jami2_3_kurs += budjet['jami__sum']
                                jami2_3_kurs_b += budjet['jami__sum']
                                jami_3_kurs_b += budjet['jami__sum']
                                yonalish_jami_3_kurs += budjet['jami__sum']
                                yonalish_jami_3_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_3_kurs += shartnoma['jami__sum']
                                jami_3_kurs_c += shartnoma['jami__sum']
                                jami2_3_kurs += shartnoma['jami__sum']
                                jami2_3_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_3_kurs += shartnoma['jami__sum']
                                yonalish_jami_3_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 4:
                            if budjet['jami__sum']:
                                jami_4_kurs += budjet['jami__sum']
                                jami_4_kurs_b += budjet['jami__sum']
                                jami2_4_kurs += budjet['jami__sum']
                                jami2_4_kurs_b += budjet['jami__sum']
                                yonalish_jami_4_kurs += budjet['jami__sum']
                                yonalish_jami_4_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs_c += shartnoma['jami__sum']
                                jami_4_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_4_kurs += shartnoma['jami__sum']
                                yonalish_jami_4_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 5:
                            if budjet['jami__sum']:
                                jami_5_kurs += budjet['jami__sum']
                                jami2_5_kurs += budjet['jami__sum']
                                jami2_5_kurs_b += budjet['jami__sum']
                                jami_5_kurs_b += budjet['jami__sum']
                                yonalish_jami_5_kurs += budjet['jami__sum']
                                yonalish_jami_5_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_5_kurs += shartnoma['jami__sum']
                                jami_5_kurs_c += shartnoma['jami__sum']
                                jami2_5_kurs += shartnoma['jami__sum']
                                jami2_5_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_5_kurs += shartnoma['jami__sum']
                                yonalish_jami_5_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 6:
                            if budjet['jami__sum']:
                                jami_6_kurs += budjet['jami__sum']
                                jami_6_kurs_b += budjet['jami__sum']
                                jami2_6_kurs += budjet['jami__sum']
                                jami2_6_kurs_b += budjet['jami__sum']
                                yonalish_jami_6_kurs += budjet['jami__sum']
                                yonalish_jami_6_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs_c += shartnoma['jami__sum']
                                jami_6_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_6_kurs += shartnoma['jami__sum']
                                yonalish_jami_6_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 7:
                            if budjet['jami__sum']:
                                jami_7_kurs += budjet['jami__sum']
                                jami2_7_kurs += budjet['jami__sum']
                                jami_7_kurs_b += budjet['jami__sum']
                                yonalish_jami_7_kurs += budjet['jami__sum']
                                yonalish_jami_7_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_7_kurs += shartnoma['jami__sum']
                                jami2_7_kurs += shartnoma['jami__sum']
                                jami_7_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_7_kurs += shartnoma['jami__sum']
                                yonalish_jami_7_kurs_c += shartnoma['jami__sum']


                if jami != 0:
                    ws.cell(row=row, column=2, value=y.name)
                    ws.cell(row=row, column=3, value='Magistir')
                    ws.cell(row=row, column=4, value=jami)
                    ws.cell(row=row, column=5, value=jami_b)
                    ws.cell(row=row, column=6, value=jami_c)
                    ws.cell(row=row, column=7, value=yonalish_jami_1_kurs)
                    ws.cell(row=row, column=8, value=yonalish_jami_1_kurs_b)
                    ws.cell(row=row, column=9, value=yonalish_jami_1_kurs_c)
                    ws.cell(row=row, column=10, value=yonalish_jami_2_kurs)
                    ws.cell(row=row, column=11, value=yonalish_jami_2_kurs_b)
                    ws.cell(row=row, column=12, value=yonalish_jami_2_kurs_c)
                    ws.cell(row=row, column=13, value=yonalish_jami_3_kurs)
                    ws.cell(row=row, column=14, value=yonalish_jami_3_kurs_b)
                    ws.cell(row=row, column=15, value=yonalish_jami_3_kurs_c)
                    ws.cell(row=row, column=16, value=yonalish_jami_4_kurs)
                    ws.cell(row=row, column=17, value=yonalish_jami_4_kurs_b)
                    ws.cell(row=row, column=18, value=yonalish_jami_4_kurs_c)
                    ws.cell(row=row, column=19, value=yonalish_jami_5_kurs)
                    ws.cell(row=row, column=20, value=yonalish_jami_5_kurs_b)
                    ws.cell(row=row, column=21, value=yonalish_jami_5_kurs_c)
                    ws.cell(row=row, column=22, value=yonalish_jami_6_kurs)
                    ws.cell(row=row, column=23, value=yonalish_jami_6_kurs_b)
                    ws.cell(row=row, column=24, value=yonalish_jami_6_kurs_c)
                    ws.cell(row=row, column=25, value=yonalish_jami_7_kurs)
                    ws.cell(row=row, column=26, value=yonalish_jami_7_kurs_b)
                    ws.cell(row=row, column=27, value=yonalish_jami_7_kurs_c)
                    row += 1
            for y in yonalish_Sirtqi:

                jami = 0
                jami_b = 0
                jami_c = 0
                yonalish_jami_1_kurs = 0
                yonalish_jami_1_kurs_b = 0
                yonalish_jami_1_kurs_c = 0
                yonalish_jami_2_kurs = 0
                yonalish_jami_2_kurs_b = 0
                yonalish_jami_2_kurs_c = 0
                yonalish_jami_3_kurs = 0
                yonalish_jami_3_kurs_b = 0
                yonalish_jami_3_kurs_c = 0
                yonalish_jami_4_kurs = 0
                yonalish_jami_4_kurs_b = 0
                yonalish_jami_4_kurs_c = 0
                yonalish_jami_5_kurs = 0
                yonalish_jami_5_kurs_b = 0
                yonalish_jami_5_kurs_c = 0
                yonalish_jami_6_kurs = 0
                yonalish_jami_6_kurs_b = 0
                yonalish_jami_6_kurs_c = 0
                yonalish_jami_7_kurs = 0
                yonalish_jami_7_kurs_b = 0
                yonalish_jami_7_kurs_c = 0
                guruhs = Guruh.objects.filter(org=org, yonalish=y, yonalish__faculty=f)
                for g in guruhs:

                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'))
                    if budjet['jami__sum']:
                        jami_full += budjet['jami__sum']
                        jami2_full += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']
                        jami += budjet['jami__sum']
                        jami_b += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']

                    if shartnoma['jami__sum']:
                        jami_full += shartnoma['jami__sum']
                        jami2_full += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                        jami += shartnoma['jami__sum']
                        jami_c += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                    if jami != 0:
                        if g.kurs == 1:
                            if budjet['jami__sum']:
                                jami_1_kurs += budjet['jami__sum']
                                jami2_1_kurs += budjet['jami__sum']
                                jami2_1_kurs_b += budjet['jami__sum']
                                jami_1_kurs_b += budjet['jami__sum']
                                yonalish_jami_1_kurs += budjet['jami__sum']
                                yonalish_jami_1_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs_c += shartnoma['jami__sum']
                                jami_1_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_1_kurs += shartnoma['jami__sum']
                                yonalish_jami_1_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 2:
                            if budjet['jami__sum']:
                                jami_2_kurs += budjet['jami__sum']
                                jami2_2_kurs += budjet['jami__sum']
                                jami2_2_kurs_b += budjet['jami__sum']
                                jami_2_kurs_b += budjet['jami__sum']
                                yonalish_jami_2_kurs += budjet['jami__sum']
                                yonalish_jami_2_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs_c += shartnoma['jami__sum']
                                jami_2_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_2_kurs += shartnoma['jami__sum']
                                yonalish_jami_2_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 3:
                            if budjet['jami__sum']:
                                jami_3_kurs += budjet['jami__sum']
                                jami2_3_kurs += budjet['jami__sum']
                                jami2_3_kurs_b += budjet['jami__sum']
                                jami_3_kurs_b += budjet['jami__sum']
                                yonalish_jami_3_kurs += budjet['jami__sum']
                                yonalish_jami_3_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_3_kurs += shartnoma['jami__sum']
                                jami_3_kurs_c += shartnoma['jami__sum']
                                jami2_3_kurs += shartnoma['jami__sum']
                                jami2_3_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_3_kurs += shartnoma['jami__sum']
                                yonalish_jami_3_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 4:
                            if budjet['jami__sum']:
                                jami_4_kurs += budjet['jami__sum']
                                jami_4_kurs_b += budjet['jami__sum']
                                jami2_4_kurs += budjet['jami__sum']
                                jami2_4_kurs_b += budjet['jami__sum']
                                yonalish_jami_4_kurs += budjet['jami__sum']
                                yonalish_jami_4_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs_c += shartnoma['jami__sum']
                                jami_4_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_4_kurs += shartnoma['jami__sum']
                                yonalish_jami_4_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 5:
                            if budjet['jami__sum']:
                                jami_5_kurs += budjet['jami__sum']
                                jami2_5_kurs += budjet['jami__sum']
                                jami2_5_kurs_b += budjet['jami__sum']
                                jami_5_kurs_b += budjet['jami__sum']
                                yonalish_jami_5_kurs += budjet['jami__sum']
                                yonalish_jami_5_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_5_kurs += shartnoma['jami__sum']
                                jami_5_kurs_c += shartnoma['jami__sum']
                                jami2_5_kurs += shartnoma['jami__sum']
                                jami2_5_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_5_kurs += shartnoma['jami__sum']
                                yonalish_jami_5_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 6:
                            if budjet['jami__sum']:
                                jami_6_kurs += budjet['jami__sum']
                                jami_6_kurs_b += budjet['jami__sum']
                                jami2_6_kurs += budjet['jami__sum']
                                jami2_6_kurs_b += budjet['jami__sum']
                                yonalish_jami_6_kurs += budjet['jami__sum']
                                yonalish_jami_6_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs_c += shartnoma['jami__sum']
                                jami_6_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_6_kurs += shartnoma['jami__sum']
                                yonalish_jami_6_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 7:
                            if budjet['jami__sum']:
                                jami_7_kurs += budjet['jami__sum']
                                jami2_7_kurs += budjet['jami__sum']
                                jami_7_kurs_b += budjet['jami__sum']
                                yonalish_jami_7_kurs += budjet['jami__sum']
                                yonalish_jami_7_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_7_kurs += shartnoma['jami__sum']
                                jami2_7_kurs += shartnoma['jami__sum']
                                jami_7_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_7_kurs += shartnoma['jami__sum']
                                yonalish_jami_7_kurs_c += shartnoma['jami__sum']



                if jami != 0:
                    ws.cell(row=row, column=2, value=y.name)
                    ws.cell(row=row, column=3, value=y.turi)
                    ws.cell(row=row, column=4, value=jami)
                    ws.cell(row=row, column=5, value=jami_b)
                    ws.cell(row=row, column=6, value=jami_c)
                    ws.cell(row=row, column=7, value=yonalish_jami_1_kurs)
                    ws.cell(row=row, column=8, value=yonalish_jami_1_kurs_b)
                    ws.cell(row=row, column=9, value=yonalish_jami_1_kurs_c)
                    ws.cell(row=row, column=10, value=yonalish_jami_2_kurs)
                    ws.cell(row=row, column=11, value=yonalish_jami_2_kurs_b)
                    ws.cell(row=row, column=12, value=yonalish_jami_2_kurs_c)
                    ws.cell(row=row, column=13, value=yonalish_jami_3_kurs)
                    ws.cell(row=row, column=14, value=yonalish_jami_3_kurs_b)
                    ws.cell(row=row, column=15, value=yonalish_jami_3_kurs_c)
                    ws.cell(row=row, column=16, value=yonalish_jami_4_kurs)
                    ws.cell(row=row, column=17, value=yonalish_jami_4_kurs_b)
                    ws.cell(row=row, column=18, value=yonalish_jami_4_kurs_c)
                    ws.cell(row=row, column=19, value=yonalish_jami_5_kurs)
                    ws.cell(row=row, column=20, value=yonalish_jami_5_kurs_b)
                    ws.cell(row=row, column=21, value=yonalish_jami_5_kurs_c)
                    ws.cell(row=row, column=22, value=yonalish_jami_6_kurs)
                    ws.cell(row=row, column=23, value=yonalish_jami_6_kurs_b)
                    ws.cell(row=row, column=24, value=yonalish_jami_6_kurs_c)
                    ws.cell(row=row, column=25, value=yonalish_jami_7_kurs)
                    ws.cell(row=row, column=26, value=yonalish_jami_7_kurs_b)
                    ws.cell(row=row, column=27, value=yonalish_jami_7_kurs_c)
                    row += 1
            for y in yonalish_Masofaviy:

                jami = 0
                jami_b = 0
                jami_c = 0
                yonalish_jami_1_kurs = 0
                yonalish_jami_1_kurs_b = 0
                yonalish_jami_1_kurs_c = 0
                yonalish_jami_2_kurs = 0
                yonalish_jami_2_kurs_b = 0
                yonalish_jami_2_kurs_c = 0
                yonalish_jami_3_kurs = 0
                yonalish_jami_3_kurs_b = 0
                yonalish_jami_3_kurs_c = 0
                yonalish_jami_4_kurs = 0
                yonalish_jami_4_kurs_b = 0
                yonalish_jami_4_kurs_c = 0
                yonalish_jami_5_kurs = 0
                yonalish_jami_5_kurs_b = 0
                yonalish_jami_5_kurs_c = 0
                yonalish_jami_6_kurs = 0
                yonalish_jami_6_kurs_b = 0
                yonalish_jami_6_kurs_c = 0
                yonalish_jami_7_kurs = 0
                yonalish_jami_7_kurs_b = 0
                yonalish_jami_7_kurs_c = 0
                guruhs = Guruh.objects.filter(org=org, yonalish=y, yonalish__faculty=f)
                for g in guruhs:

                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'))
                    if budjet['jami__sum']:
                        jami_full += budjet['jami__sum']
                        jami2_full += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']
                        jami += budjet['jami__sum']
                        jami_b += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']

                    if shartnoma['jami__sum']:
                        jami_full += shartnoma['jami__sum']
                        jami2_full += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                        jami += shartnoma['jami__sum']
                        jami_c += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                    if jami != 0:
                        if g.kurs == 1:
                            if budjet['jami__sum']:
                                jami_1_kurs += budjet['jami__sum']
                                jami2_1_kurs += budjet['jami__sum']
                                jami2_1_kurs_b += budjet['jami__sum']
                                jami_1_kurs_b += budjet['jami__sum']
                                yonalish_jami_1_kurs += budjet['jami__sum']
                                yonalish_jami_1_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs_c += shartnoma['jami__sum']
                                jami_1_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_1_kurs += shartnoma['jami__sum']
                                yonalish_jami_1_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 2:
                            if budjet['jami__sum']:
                                jami_2_kurs += budjet['jami__sum']
                                jami2_2_kurs += budjet['jami__sum']
                                jami2_2_kurs_b += budjet['jami__sum']
                                jami_2_kurs_b += budjet['jami__sum']
                                yonalish_jami_2_kurs += budjet['jami__sum']
                                yonalish_jami_2_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs_c += shartnoma['jami__sum']
                                jami_2_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_2_kurs += shartnoma['jami__sum']
                                yonalish_jami_2_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 3:
                            if budjet['jami__sum']:
                                jami_3_kurs += budjet['jami__sum']
                                jami2_3_kurs += budjet['jami__sum']
                                jami2_3_kurs_b += budjet['jami__sum']
                                jami_3_kurs_b += budjet['jami__sum']
                                yonalish_jami_3_kurs += budjet['jami__sum']
                                yonalish_jami_3_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_3_kurs += shartnoma['jami__sum']
                                jami_3_kurs_c += shartnoma['jami__sum']
                                jami2_3_kurs += shartnoma['jami__sum']
                                jami2_3_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_3_kurs += shartnoma['jami__sum']
                                yonalish_jami_3_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 4:
                            if budjet['jami__sum']:
                                jami_4_kurs += budjet['jami__sum']
                                jami_4_kurs_b += budjet['jami__sum']
                                jami2_4_kurs += budjet['jami__sum']
                                jami2_4_kurs_b += budjet['jami__sum']
                                yonalish_jami_4_kurs += budjet['jami__sum']
                                yonalish_jami_4_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs_c += shartnoma['jami__sum']
                                jami_4_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_4_kurs += shartnoma['jami__sum']
                                yonalish_jami_4_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 5:
                            if budjet['jami__sum']:
                                jami_5_kurs += budjet['jami__sum']
                                jami2_5_kurs += budjet['jami__sum']
                                jami2_5_kurs_b += budjet['jami__sum']
                                jami_5_kurs_b += budjet['jami__sum']
                                yonalish_jami_5_kurs += budjet['jami__sum']
                                yonalish_jami_5_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_5_kurs += shartnoma['jami__sum']
                                jami_5_kurs_c += shartnoma['jami__sum']
                                jami2_5_kurs += shartnoma['jami__sum']
                                jami2_5_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_5_kurs += shartnoma['jami__sum']
                                yonalish_jami_5_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 6:
                            if budjet['jami__sum']:
                                jami_6_kurs += budjet['jami__sum']
                                jami_6_kurs_b += budjet['jami__sum']
                                jami2_6_kurs += budjet['jami__sum']
                                jami2_6_kurs_b += budjet['jami__sum']
                                yonalish_jami_6_kurs += budjet['jami__sum']
                                yonalish_jami_6_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs_c += shartnoma['jami__sum']
                                jami_6_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_6_kurs += shartnoma['jami__sum']
                                yonalish_jami_6_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 7:
                            if budjet['jami__sum']:
                                jami_7_kurs += budjet['jami__sum']
                                jami2_7_kurs += budjet['jami__sum']
                                jami_7_kurs_b += budjet['jami__sum']
                                yonalish_jami_7_kurs += budjet['jami__sum']
                                yonalish_jami_7_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_7_kurs += shartnoma['jami__sum']
                                jami2_7_kurs += shartnoma['jami__sum']
                                jami_7_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_7_kurs += shartnoma['jami__sum']
                                yonalish_jami_7_kurs_c += shartnoma['jami__sum']


                if jami != 0:

                    ws.cell(row=row, column=2, value=y.name)
                    ws.cell(row=row, column=3, value=y.turi)
                    ws.cell(row=row, column=4, value=jami)
                    ws.cell(row=row, column=5, value=jami_b)
                    ws.cell(row=row, column=6, value=jami_c)
                    ws.cell(row=row, column=7, value=yonalish_jami_1_kurs)
                    ws.cell(row=row, column=8, value=yonalish_jami_1_kurs_b)
                    ws.cell(row=row, column=9, value=yonalish_jami_1_kurs_c)
                    ws.cell(row=row, column=10, value=yonalish_jami_2_kurs)
                    ws.cell(row=row, column=11, value=yonalish_jami_2_kurs_b)
                    ws.cell(row=row, column=12, value=yonalish_jami_2_kurs_c)
                    ws.cell(row=row, column=13, value=yonalish_jami_3_kurs)
                    ws.cell(row=row, column=14, value=yonalish_jami_3_kurs_b)
                    ws.cell(row=row, column=15, value=yonalish_jami_3_kurs_c)
                    ws.cell(row=row, column=16, value=yonalish_jami_4_kurs)
                    ws.cell(row=row, column=17, value=yonalish_jami_4_kurs_b)
                    ws.cell(row=row, column=18, value=yonalish_jami_4_kurs_c)
                    ws.cell(row=row, column=19, value=yonalish_jami_5_kurs)
                    ws.cell(row=row, column=20, value=yonalish_jami_5_kurs_b)
                    ws.cell(row=row, column=21, value=yonalish_jami_5_kurs_c)
                    ws.cell(row=row, column=22, value=yonalish_jami_6_kurs)
                    ws.cell(row=row, column=23, value=yonalish_jami_6_kurs_b)
                    ws.cell(row=row, column=24, value=yonalish_jami_6_kurs_c)
                    ws.cell(row=row, column=25, value=yonalish_jami_7_kurs)
                    ws.cell(row=row, column=26, value=yonalish_jami_7_kurs_b)
                    ws.cell(row=row, column=27, value=yonalish_jami_7_kurs_c)
                    row += 1
            for y in yonalish_Doktorantura:

                jami = 0
                jami_b = 0
                jami_c = 0
                yonalish_jami_1_kurs = 0
                yonalish_jami_1_kurs_b = 0
                yonalish_jami_1_kurs_c = 0
                yonalish_jami_2_kurs = 0
                yonalish_jami_2_kurs_b = 0
                yonalish_jami_2_kurs_c = 0
                yonalish_jami_3_kurs = 0
                yonalish_jami_3_kurs_b = 0
                yonalish_jami_3_kurs_c = 0
                yonalish_jami_4_kurs = 0
                yonalish_jami_4_kurs_b = 0
                yonalish_jami_4_kurs_c = 0
                yonalish_jami_5_kurs = 0
                yonalish_jami_5_kurs_b = 0
                yonalish_jami_5_kurs_c = 0
                yonalish_jami_6_kurs = 0
                yonalish_jami_6_kurs_b = 0
                yonalish_jami_6_kurs_c = 0
                yonalish_jami_7_kurs = 0
                yonalish_jami_7_kurs_b = 0
                yonalish_jami_7_kurs_c = 0
                guruhs = Guruh.objects.filter(org=org, yonalish=y, yonalish__faculty=f)
                for g in guruhs:

                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'))
                    if budjet['jami__sum']:
                        jami_full += budjet['jami__sum']
                        jami2_full += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']
                        jami += budjet['jami__sum']
                        jami_b += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']

                    if shartnoma['jami__sum']:
                        jami_full += shartnoma['jami__sum']
                        jami2_full += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                        jami += shartnoma['jami__sum']
                        jami_c += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                    if jami != 0:
                        if g.kurs == 1:
                            if budjet['jami__sum']:
                                jami_1_kurs += budjet['jami__sum']
                                jami2_1_kurs += budjet['jami__sum']
                                jami2_1_kurs_b += budjet['jami__sum']
                                jami_1_kurs_b += budjet['jami__sum']
                                yonalish_jami_1_kurs += budjet['jami__sum']
                                yonalish_jami_1_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs_c += shartnoma['jami__sum']
                                jami_1_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_1_kurs += shartnoma['jami__sum']
                                yonalish_jami_1_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 2:
                            if budjet['jami__sum']:
                                jami_2_kurs += budjet['jami__sum']
                                jami2_2_kurs += budjet['jami__sum']
                                jami2_2_kurs_b += budjet['jami__sum']
                                jami_2_kurs_b += budjet['jami__sum']
                                yonalish_jami_2_kurs += budjet['jami__sum']
                                yonalish_jami_2_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs_c += shartnoma['jami__sum']
                                jami_2_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_2_kurs += shartnoma['jami__sum']
                                yonalish_jami_2_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 3:
                            if budjet['jami__sum']:
                                jami_3_kurs += budjet['jami__sum']
                                jami2_3_kurs += budjet['jami__sum']
                                jami2_3_kurs_b += budjet['jami__sum']
                                jami_3_kurs_b += budjet['jami__sum']
                                yonalish_jami_3_kurs += budjet['jami__sum']
                                yonalish_jami_3_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_3_kurs += shartnoma['jami__sum']
                                jami_3_kurs_c += shartnoma['jami__sum']
                                jami2_3_kurs += shartnoma['jami__sum']
                                jami2_3_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_3_kurs += shartnoma['jami__sum']
                                yonalish_jami_3_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 4:
                            if budjet['jami__sum']:
                                jami_4_kurs += budjet['jami__sum']
                                jami_4_kurs_b += budjet['jami__sum']
                                jami2_4_kurs += budjet['jami__sum']
                                jami2_4_kurs_b += budjet['jami__sum']
                                yonalish_jami_4_kurs += budjet['jami__sum']
                                yonalish_jami_4_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs_c += shartnoma['jami__sum']
                                jami_4_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_4_kurs += shartnoma['jami__sum']
                                yonalish_jami_4_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 5:
                            if budjet['jami__sum']:
                                jami_5_kurs += budjet['jami__sum']
                                jami2_5_kurs += budjet['jami__sum']
                                jami2_5_kurs_b += budjet['jami__sum']
                                jami_5_kurs_b += budjet['jami__sum']
                                yonalish_jami_5_kurs += budjet['jami__sum']
                                yonalish_jami_5_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_5_kurs += shartnoma['jami__sum']
                                jami_5_kurs_c += shartnoma['jami__sum']
                                jami2_5_kurs += shartnoma['jami__sum']
                                jami2_5_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_5_kurs += shartnoma['jami__sum']
                                yonalish_jami_5_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 6:
                            if budjet['jami__sum']:
                                jami_6_kurs += budjet['jami__sum']
                                jami_6_kurs_b += budjet['jami__sum']
                                jami2_6_kurs += budjet['jami__sum']
                                jami2_6_kurs_b += budjet['jami__sum']
                                yonalish_jami_6_kurs += budjet['jami__sum']
                                yonalish_jami_6_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs_c += shartnoma['jami__sum']
                                jami_6_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_6_kurs += shartnoma['jami__sum']
                                yonalish_jami_6_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 7:
                            if budjet['jami__sum']:
                                jami_7_kurs += budjet['jami__sum']
                                jami2_7_kurs += budjet['jami__sum']
                                jami_7_kurs_b += budjet['jami__sum']
                                yonalish_jami_7_kurs += budjet['jami__sum']
                                yonalish_jami_7_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_7_kurs += shartnoma['jami__sum']
                                jami2_7_kurs += shartnoma['jami__sum']
                                jami_7_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_7_kurs += shartnoma['jami__sum']
                                yonalish_jami_7_kurs_c += shartnoma['jami__sum']


                if jami != 0:
                    ws.cell(row=row, column=2, value=y.name)
                    ws.cell(row=row, column=3, value=y.turi)
                    ws.cell(row=row, column=4, value=jami)
                    ws.cell(row=row, column=5, value=jami_b)
                    ws.cell(row=row, column=6, value=jami_c)
                    ws.cell(row=row, column=7, value=yonalish_jami_1_kurs)
                    ws.cell(row=row, column=8, value=yonalish_jami_1_kurs_b)
                    ws.cell(row=row, column=9, value=yonalish_jami_1_kurs_c)
                    ws.cell(row=row, column=10, value=yonalish_jami_2_kurs)
                    ws.cell(row=row, column=11, value=yonalish_jami_2_kurs_b)
                    ws.cell(row=row, column=12, value=yonalish_jami_2_kurs_c)
                    ws.cell(row=row, column=13, value=yonalish_jami_3_kurs)
                    ws.cell(row=row, column=14, value=yonalish_jami_3_kurs_b)
                    ws.cell(row=row, column=15, value=yonalish_jami_3_kurs_c)
                    ws.cell(row=row, column=16, value=yonalish_jami_4_kurs)
                    ws.cell(row=row, column=17, value=yonalish_jami_4_kurs_b)
                    ws.cell(row=row, column=18, value=yonalish_jami_4_kurs_c)
                    ws.cell(row=row, column=19, value=yonalish_jami_5_kurs)
                    ws.cell(row=row, column=20, value=yonalish_jami_5_kurs_b)
                    ws.cell(row=row, column=21, value=yonalish_jami_5_kurs_c)
                    ws.cell(row=row, column=22, value=yonalish_jami_6_kurs)
                    ws.cell(row=row, column=23, value=yonalish_jami_6_kurs_b)
                    ws.cell(row=row, column=24, value=yonalish_jami_6_kurs_c)
                    ws.cell(row=row, column=25, value=yonalish_jami_7_kurs)
                    ws.cell(row=row, column=26, value=yonalish_jami_7_kurs_b)
                    ws.cell(row=row, column=27, value=yonalish_jami_7_kurs_c)
                    row += 1
            for y in yonalish_ikkinchi_talim:

                jami = 0
                jami_b = 0
                jami_c = 0
                yonalish_jami_1_kurs = 0
                yonalish_jami_1_kurs_b = 0
                yonalish_jami_1_kurs_c = 0
                yonalish_jami_2_kurs = 0
                yonalish_jami_2_kurs_b = 0
                yonalish_jami_2_kurs_c = 0
                yonalish_jami_3_kurs = 0
                yonalish_jami_3_kurs_b = 0
                yonalish_jami_3_kurs_c = 0
                yonalish_jami_4_kurs = 0
                yonalish_jami_4_kurs_b = 0
                yonalish_jami_4_kurs_c = 0
                yonalish_jami_5_kurs = 0
                yonalish_jami_5_kurs_b = 0
                yonalish_jami_5_kurs_c = 0
                yonalish_jami_6_kurs = 0
                yonalish_jami_6_kurs_b = 0
                yonalish_jami_6_kurs_c = 0
                yonalish_jami_7_kurs = 0
                yonalish_jami_7_kurs_b = 0
                yonalish_jami_7_kurs_c = 0
                guruhs = Guruh.objects.filter(org=org, yonalish=y, yonalish__faculty=f)
                for g in guruhs:

                    budjet = Budjet.objects.filter(guruhi=g).aggregate(Sum('jami'))
                    shartnoma = Shartnoma.objects.filter(guruh=g).aggregate(Sum('jami'))
                    if budjet['jami__sum']:
                        jami_full += budjet['jami__sum']
                        jami2_full += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']
                        jami += budjet['jami__sum']
                        jami_b += budjet['jami__sum']
                        jami2_b += budjet['jami__sum']

                    if shartnoma['jami__sum']:
                        jami_full += shartnoma['jami__sum']
                        jami2_full += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                        jami += shartnoma['jami__sum']
                        jami_c += shartnoma['jami__sum']
                        jami2_c += shartnoma['jami__sum']
                    if jami != 0:
                        if g.kurs == 1:
                            if budjet['jami__sum']:
                                jami_1_kurs += budjet['jami__sum']
                                jami2_1_kurs += budjet['jami__sum']
                                jami2_1_kurs_b += budjet['jami__sum']
                                jami_1_kurs_b += budjet['jami__sum']
                                yonalish_jami_1_kurs += budjet['jami__sum']
                                yonalish_jami_1_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs += shartnoma['jami__sum']
                                jami2_1_kurs_c += shartnoma['jami__sum']
                                jami_1_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_1_kurs += shartnoma['jami__sum']
                                yonalish_jami_1_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 2:
                            if budjet['jami__sum']:
                                jami_2_kurs += budjet['jami__sum']
                                jami2_2_kurs += budjet['jami__sum']
                                jami2_2_kurs_b += budjet['jami__sum']
                                jami_2_kurs_b += budjet['jami__sum']
                                yonalish_jami_2_kurs += budjet['jami__sum']
                                yonalish_jami_2_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs += shartnoma['jami__sum']
                                jami2_2_kurs_c += shartnoma['jami__sum']
                                jami_2_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_2_kurs += shartnoma['jami__sum']
                                yonalish_jami_2_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 3:
                            if budjet['jami__sum']:
                                jami_3_kurs += budjet['jami__sum']
                                jami2_3_kurs += budjet['jami__sum']
                                jami2_3_kurs_b += budjet['jami__sum']
                                jami_3_kurs_b += budjet['jami__sum']
                                yonalish_jami_3_kurs += budjet['jami__sum']
                                yonalish_jami_3_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_3_kurs += shartnoma['jami__sum']
                                jami_3_kurs_c += shartnoma['jami__sum']
                                jami2_3_kurs += shartnoma['jami__sum']
                                jami2_3_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_3_kurs += shartnoma['jami__sum']
                                yonalish_jami_3_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 4:
                            if budjet['jami__sum']:
                                jami_4_kurs += budjet['jami__sum']
                                jami_4_kurs_b += budjet['jami__sum']
                                jami2_4_kurs += budjet['jami__sum']
                                jami2_4_kurs_b += budjet['jami__sum']
                                yonalish_jami_4_kurs += budjet['jami__sum']
                                yonalish_jami_4_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs += shartnoma['jami__sum']
                                jami2_4_kurs_c += shartnoma['jami__sum']
                                jami_4_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_4_kurs += shartnoma['jami__sum']
                                yonalish_jami_4_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 5:
                            if budjet['jami__sum']:
                                jami_5_kurs += budjet['jami__sum']
                                jami2_5_kurs += budjet['jami__sum']
                                jami2_5_kurs_b += budjet['jami__sum']
                                jami_5_kurs_b += budjet['jami__sum']
                                yonalish_jami_5_kurs += budjet['jami__sum']
                                yonalish_jami_5_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_5_kurs += shartnoma['jami__sum']
                                jami_5_kurs_c += shartnoma['jami__sum']
                                jami2_5_kurs += shartnoma['jami__sum']
                                jami2_5_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_5_kurs += shartnoma['jami__sum']
                                yonalish_jami_5_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 6:
                            if budjet['jami__sum']:
                                jami_6_kurs += budjet['jami__sum']
                                jami_6_kurs_b += budjet['jami__sum']
                                jami2_6_kurs += budjet['jami__sum']
                                jami2_6_kurs_b += budjet['jami__sum']
                                yonalish_jami_6_kurs += budjet['jami__sum']
                                yonalish_jami_6_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs += shartnoma['jami__sum']
                                jami2_6_kurs_c += shartnoma['jami__sum']
                                jami_6_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_6_kurs += shartnoma['jami__sum']
                                yonalish_jami_6_kurs_c += shartnoma['jami__sum']
                        if g.kurs == 7:
                            if budjet['jami__sum']:
                                jami_7_kurs += budjet['jami__sum']
                                jami2_7_kurs += budjet['jami__sum']
                                jami_7_kurs_b += budjet['jami__sum']
                                yonalish_jami_7_kurs += budjet['jami__sum']
                                yonalish_jami_7_kurs_b += budjet['jami__sum']
                            if shartnoma['jami__sum']:
                                jami_7_kurs += shartnoma['jami__sum']
                                jami2_7_kurs += shartnoma['jami__sum']
                                jami_7_kurs_c += shartnoma['jami__sum']
                                yonalish_jami_7_kurs += shartnoma['jami__sum']
                                yonalish_jami_7_kurs_c += shartnoma['jami__sum']


                if jami != 0:
                    ws.cell(row=row, column=2, value=y.name)
                    ws.cell(row=row, column=3, value=y.turi)
                    ws.cell(row=row, column=4, value=jami)
                    ws.cell(row=row, column=5, value=jami_b)
                    ws.cell(row=row, column=6, value=jami_c)
                    ws.cell(row=row, column=7, value=yonalish_jami_1_kurs)
                    ws.cell(row=row, column=8, value=yonalish_jami_1_kurs_b)
                    ws.cell(row=row, column=9, value=yonalish_jami_1_kurs_c)
                    ws.cell(row=row, column=10, value=yonalish_jami_2_kurs)
                    ws.cell(row=row, column=11, value=yonalish_jami_2_kurs_b)
                    ws.cell(row=row, column=12, value=yonalish_jami_2_kurs_c)
                    ws.cell(row=row, column=13, value=yonalish_jami_3_kurs)
                    ws.cell(row=row, column=14, value=yonalish_jami_3_kurs_b)
                    ws.cell(row=row, column=15, value=yonalish_jami_3_kurs_c)
                    ws.cell(row=row, column=16, value=yonalish_jami_4_kurs)
                    ws.cell(row=row, column=17, value=yonalish_jami_4_kurs_b)
                    ws.cell(row=row, column=18, value=yonalish_jami_4_kurs_c)
                    ws.cell(row=row, column=19, value=yonalish_jami_5_kurs)
                    ws.cell(row=row, column=20, value=yonalish_jami_5_kurs_b)
                    ws.cell(row=row, column=21, value=yonalish_jami_5_kurs_c)
                    ws.cell(row=row, column=22, value=yonalish_jami_6_kurs)
                    ws.cell(row=row, column=23, value=yonalish_jami_6_kurs_b)
                    ws.cell(row=row, column=24, value=yonalish_jami_6_kurs_c)
                    ws.cell(row=row, column=25, value=yonalish_jami_7_kurs)
                    ws.cell(row=row, column=26, value=yonalish_jami_7_kurs_b)
                    ws.cell(row=row, column=27, value=yonalish_jami_7_kurs_c)
                    row += 1
            row += 1

            ws.cell(row=row, column=2, value=f'{f.name} jami')
            ws.cell(row=row, column=4, value=jami_full)
            ws.cell(row=row, column=5, value=jami_b)
            ws.cell(row=row, column=6, value=jami_c)
            ws.cell(row=row, column=7, value=jami_1_kurs)
            ws.cell(row=row, column=8, value=jami_1_kurs_b)
            ws.cell(row=row, column=9, value=jami_1_kurs_c)
            ws.cell(row=row, column=10, value=jami_2_kurs)
            ws.cell(row=row, column=11, value=jami_2_kurs_b)
            ws.cell(row=row, column=12, value=jami_2_kurs_c)
            ws.cell(row=row, column=13, value=jami_3_kurs)
            ws.cell(row=row, column=14, value=jami_3_kurs_b)
            ws.cell(row=row, column=15, value=jami_3_kurs_c)
            ws.cell(row=row, column=16, value=jami_4_kurs)
            ws.cell(row=row, column=17, value=jami_4_kurs_b)
            ws.cell(row=row, column=18, value=jami_4_kurs_c)
            ws.cell(row=row, column=19, value=jami_5_kurs)

            ws.cell(row=row, column=20, value=jami_5_kurs_b)
            ws.cell(row=row, column=21, value=jami_5_kurs_c)
            ws.cell(row=row, column=22, value=jami_6_kurs)
            ws.cell(row=row, column=23, value=jami_6_kurs_b)
            ws.cell(row=row, column=24, value=jami_6_kurs_c)
            ws.cell(row=row, column=25, value=jami_7_kurs)
            ws.cell(row=row, column=26, value=jami_7_kurs_b)
            ws.cell(row=row, column=27, value=jami_7_kurs_c)
            row += 1

    wb.save(output4)
    output4.seek(0)
    return output4


